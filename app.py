# app.py  ✅ FULL COPY-PASTE FILE (WITH COALESCE KEY FIX FOR SEP FILE)
# FIXES:
# 1) KEY cleaning: if key column is text (loan account numbers), keep alphanumerics (no digit-only stripping)
# 2) Only-in vs mismatch logic corrected (only-in never counted as mismatch)
# 3) Remove "Grand Total ignore match" row
# 4) Diff always; TRUE/FALSE match only for TEXT focus columns
# 5) Step-2 optimized by aggregating only needed columns
#
# ✅ NEW FIX ADDED NOW:
# D) ✅ COALESCE KEY (Sep file issue): if "Lender Loan Account Number" is blank,
#    automatically use "loan account number" (or "Loan Account Number") as key.
#    Works both in 2.5 summary and in final Run.
#
# ✅ NEW FIX ADDED (LOOP FIX):
# E) ✅ Prevent recomputation loop on Download clicks:
#    - Cache RUN outputs in session_state
#    - Skip Steps 1–6 if cached
#    - Reset RUN cache when uploads change (Patch-3)

# ── AUTH GATE ──
import gspread
from google.oauth2.service_account import Credentials
from datetime import date
import streamlit as st

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
creds  = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
gc     = gspread.authorize(creds)
sheet  = gc.open_by_key(st.secrets["SHEET_ID"]).sheet1
rows   = sheet.get_all_records()
user_email = st.query_params.get("email", "")
user_pin   = st.query_params.get("pin", "")

if not user_email:
    st.markdown("""...""")
    st.stop()
    st.error("❌ Access denied. Contact aarohisharma5000@gmail.com for your link.")
    st.stop()
user_row = next((r for r in rows if str(r.get("email","")).strip().lower() == user_email.strip().lower()), None)
if not user_row:
    st.error(f"❌ {user_email} is not subscribed.")
    st.stop()

expiry = date.fromisoformat(str(user_row["expiry_date"]))
if expiry < date.today():
    st.error(f"⏰ Subscription expired on {expiry}. Please renew.")
    st.stop()

# ── PIN CHECK ──
correct_pin = str(user_row.get("pin","")).strip()
if correct_pin and user_pin != correct_pin:
    st.markdown("""
    <div style="text-align:center; padding:3rem 2rem;">
        <div style="font-size:2rem; margin-bottom:1rem;">🔐</div>
        <h2 style="color:#0a1628;">Enter Your PIN</h2>
        <p style="color:#666; margin-bottom:1.5rem;">Enter the 4-digit PIN sent to you with your subscription link.</p>
    </div>
    """, unsafe_allow_html=True)
    pin_input = st.text_input("Enter PIN", type="password", max_chars=6, placeholder="Enter your PIN...")
    if st.button("🔓 Access Tool", type="primary"):
        if pin_input.strip() == correct_pin:
            st.query_params["pin"] = pin_input.strip()
            st.rerun()
        else:
            st.error("❌ Incorrect PIN. Please check your subscription email.")
    st.stop()
# ── PIN PASSED ──    
plan = str(user_row.get("plan","free")).strip().lower()
ROW_LIMIT = {"free":100,"starter":250000,"pro":9999999}.get(plan,100)
st.session_state["user_plan"] = plan
st.session_state["row_limit"] = ROW_LIMIT
# ── AUTH PASSED ──

import io
import time
import re
import pandas as pd
import streamlit as st

from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

import zipfile
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4, landscape

st.set_page_config(page_title="Reco Tool", layout="wide")

# ✅ Help moved to Launcher
show_help = False

# ── NAVY HEADER ──
import streamlit as st
_plan = st.session_state.get("user_plan", "")
_email = st.session_state.get("user_email", "")
_plan_label = {"free":"🆓 Free — 100 rows","starter":"⭐ Starter Plan","pro":"🏆 Pro Plan"}.get(_plan,"")
st.markdown(f"""
<style>
  .reco-header {{
    display:flex; align-items:center; justify-content:space-between;
    background:linear-gradient(135deg,#0a1628,#0f2040);
    border-radius:16px; padding:1rem 2rem; margin-bottom:1rem;
    border:1px solid rgba(0,212,255,0.2);
  }}
  .reco-logo {{ font-size:1.4rem; font-weight:900; color:#00d4ff; }}
  .reco-logo span {{ color:white; }}
  .reco-user {{ color:rgba(255,255,255,0.6); font-size:0.8rem; margin-top:0.2rem; }}
  .reco-plan {{
    background:rgba(0,212,255,0.15); border:1px solid rgba(0,212,255,0.3);
    color:#00d4ff; padding:0.3rem 1rem; border-radius:50px;
    font-size:0.82rem; font-weight:700;
  }}
</style>
<div class="reco-header">
  <div>
    <div class="reco-logo">Reco<span>Suite</span></div>
    <div class="reco-user">👤 {st.query_params.get("email", "")}</div>
  </div>
  {"<div class='reco-plan'>" + _plan_label + "</div>" if _plan_label else ""}
</div>
""", unsafe_allow_html=True)


# ============================================================
# 🔹 App Meta + Layout (PASTE ONCE, NEVER TOUCH AGAIN)
# ============================================================

APP_VERSION  = "v1.6"
LAST_UPDATED = "2026-03-23"
COMPANY_NAME = "RecoSuite"

# ---------- Title ----------
st.markdown(f"#### 📌 Reconciliation Tool (Multi-File Upload) &nbsp; `{APP_VERSION}`")

# ---------- Version Banner ----------
st.info(
    f"🟢 **Version:** {APP_VERSION}   |   "
    f"🕒 **Last Updated:** {LAST_UPDATED}   |   "
    f"☁️ Powered by RecoSuite Cloud",
    icon="ℹ️"
)


# ✅ Help guide removed from Tool 1 — now available in Launcher Help section
pass

# ---------- Page Padding (prevents footer overlap) ----------
st.markdown("""
<style>
.block-container {
    padding-bottom: 65px;   /* space for footer */
}
</style>
""", unsafe_allow_html=True)

# ---------- Fixed Footer ----------
st.markdown(
    f"""
    <style>
        .footer {{
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            background: #ffffff;
            border-top: 1px solid #e6e6e6;
            padding: 8px 18px;
            font-size: 12px;
            color: #666;
            z-index: 9999;
        }}
        .footer b {{
            color: #444;
        }}
    </style>

    <div class="footer">
        <b>RecoSuite</b> · Smart Reconciliation Tool · Support: aarohisharma5000@gmail.com · <a href="https://wa.me/919818799197" style="color:#0d6efd;">WhatsApp</a>
    </div>
    """,
    unsafe_allow_html=True
)

# ============================================================
st.caption("Upload multiple files on both sides (CSV/XLSX/XLSB). Each side is combined, then reco runs once.")

# -----------------------------
# Helpers
# -----------------------------
def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize headers to avoid invisible-space / newline / case issues."""
    if df is None or df.empty:
        return df
    df = df.copy()

    def _norm(c):
        c = str(c)
        c = c.replace("\u00A0", " ")   # NBSP
        c = c.replace("\u200B", "")    # zero-width
        c = c.replace("\n", " ").replace("\r", " ")
        c = c.strip()
        c = re.sub(r"\s+", " ", c)    # collapse multi spaces
        return c

    df.columns = [_norm(c) for c in df.columns]

    # ✅ FIX: dedup column names AFTER normalizing
    seen = {}
    new_cols = []
    for c in df.columns:
        if c not in seen:
            seen[c] = 1
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}__{seen[c]}")
    df.columns = new_cols

    return df


def _dedup_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    ✅ FIX 1: Make duplicate column names unique.
    e.g. ['BCF %', 'BCF %'] → ['BCF %', 'BCF %__2']
    """
    seen = {}
    new_cols = []
    for c in df.columns:
        if c not in seen:
            seen[c] = 1
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}__{seen[c]}")
    df.columns = new_cols
    return df


def _fix_mixed_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    ✅ FIX 2: Convert columns with mixed types (e.g. datetime.time, datetime.date)
    to plain strings so pyarrow / Streamlit dataframe display doesn't crash.
    """
    for col in df.columns:
        if df[col].dtype == object:
            try:
                # Check if any value is a datetime.time or datetime.date object
                sample = df[col].dropna().head(100)
                has_time = any(
                    hasattr(v, 'hour') or hasattr(v, 'year')
                    for v in sample
                    if not isinstance(v, str)
                )
                if has_time:
                    df[col] = df[col].astype(str)
            except Exception:
                pass
        # Also convert any remaining problematic object columns to str safely
        try:
            import pyarrow as pa
            pa.array(df[col].values)
        except Exception:
            df[col] = df[col].astype(str)
    return df


def read_file_once(uploaded_file) -> pd.DataFrame:
    """
    Reads CSV/XLSX/XLS/XLSB.
    For Excel files: reads ALL sheets and picks the sheet with MAX rows.
    Adds _SHEET column so summary can show which sheet was picked.
    """
    name = uploaded_file.name.lower()
    b = uploaded_file.getvalue()
    bio = io.BytesIO(b)

    if name.endswith(".csv"):
        df = pd.read_csv(bio)
        df = _dedup_columns(df)      # ✅ FIX 1
        df = _fix_mixed_types(df)    # ✅ FIX 2
        df["_SHEET"] = "CSV"
        return df

    elif name.endswith((".xlsx", ".xls")):
        sheets = pd.read_excel(bio, sheet_name=None)
        best_sheet, best_df, best_rows = None, None, -1
        for sh, dfx in sheets.items():
            if dfx is None:
                continue
            rows = len(dfx)
            if rows > best_rows:
                best_rows = rows
                best_sheet = sh
                best_df = dfx
        best_df = best_df if best_df is not None else pd.DataFrame()
        best_df = _dedup_columns(best_df)    # ✅ FIX 1
        best_df = _fix_mixed_types(best_df)  # ✅ FIX 2
        best_df["_SHEET"] = str(best_sheet) if best_sheet is not None else "UNKNOWN"
        return best_df

    elif name.endswith(".xlsb"):
        sheets = pd.read_excel(bio, engine="pyxlsb", sheet_name=None)
        best_sheet, best_df, best_rows = None, None, -1
        for sh, dfx in sheets.items():
            if dfx is None:
                continue
            rows = len(dfx)
            if rows > best_rows:
                best_rows = rows
                best_sheet = sh
                best_df = dfx
        best_df = best_df if best_df is not None else pd.DataFrame()
        best_df = _dedup_columns(best_df)    # ✅ FIX 1
        best_df = _fix_mixed_types(best_df)  # ✅ FIX 2
        best_df["_SHEET"] = str(best_sheet) if best_sheet is not None else "UNKNOWN"
        return best_df

    else:
        raise ValueError("Unsupported file type. Upload CSV / XLSX / XLSB.")
    
def read_many(files):
    """
    Combine many files into ONE dataframe (append rows).
    Returns: (combined_df, file_summary_df)
    file_summary_df columns: File, Sheet, Rows, Size_KB, Time_Sec
    Adds GRAND TOTAL row at bottom.
    Shows progress + time while reading.
    """
    dfs = []
    meta = []

    prog = st.progress(0)
    status = st.empty()

    total_files = len(files) if files else 0

    for i, uf in enumerate(files, start=1):
        t0 = time.time()
        status.info(f"Reading file {i}/{total_files}: {uf.name}")

        df = normalize_cols(read_file_once(uf))
        df["_SOURCE_FILE"] = uf.name
        dfs.append(df)

        elapsed = round(time.time() - t0, 2)

        meta.append({
            "File": uf.name,
            "Sheet": df["_SHEET"].iloc[0] if ("_SHEET" in df.columns and len(df) > 0) else "UNKNOWN",
            "Rows": int(len(df)),
            "Size_KB": round(len(uf.getvalue()) / 1024, 1),
            "Time_Sec": elapsed
        })

        prog.progress(int(i * 100 / max(1, total_files)))

    status.success("✅ File reading completed")
    prog.progress(100)

    combined = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    summary = pd.DataFrame(meta) if meta else pd.DataFrame(columns=["File", "Sheet", "Rows", "Size_KB", "Time_Sec"])

    if not summary.empty:
        summary = summary.sort_values(["Rows", "File"], ascending=[False, True])

        grand = {
            "File": "GRAND TOTAL",
            "Sheet": "",
            "Rows": int(summary["Rows"].sum()),
            "Size_KB": round(float(summary["Size_KB"].sum()), 1),
            "Time_Sec": round(float(summary["Time_Sec"].sum()), 2)
        }
        summary = pd.concat([summary, pd.DataFrame([grand])], ignore_index=True)

    return combined, summary

# ✅ treat '-' etc as blank
def is_blank(x) -> bool:
    if pd.isna(x):
        return True
    if isinstance(x, str):
        t = x.strip().lower()
        if t == "":
            return True
        if t in {"-", "--", "na", "n/a", "null", "none"}:
            return True
    return False

def uniq_list(seq):
    seen = set()
    out = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

# -----------------------------
# Decide TEXT vs NUMERIC columns
# -----------------------------
TEXT_NAME_KEYWORDS = [
    "loan account", "lender loan", "account number", "lan", "loan id", "customer id", "id"
]

def is_text_by_name(col: str) -> bool:
    lc = str(col).strip().lower()
    return any(k in lc for k in TEXT_NAME_KEYWORDS)

def is_numeric_column(df_both: pd.DataFrame, col: str) -> bool:
    if is_text_by_name(col):
        return False

    c1 = f"{col}_f1"
    c2 = f"{col}_f2"
    if c1 not in df_both.columns or c2 not in df_both.columns:
        return False

    s1 = df_both[c1]
    s2 = df_both[c2]

    nb1 = ~s1.apply(is_blank)
    nb2 = ~s2.apply(is_blank)

    n1 = pd.to_numeric(s1, errors="coerce")
    n2 = pd.to_numeric(s2, errors="coerce")

    r1 = (n1.notna() & nb1).sum() / max(1, nb1.sum())
    r2 = (n2.notna() & nb2).sum() / max(1, nb2.sum())

    return (r1 >= 0.80) and (r2 >= 0.80)

# -----------------------------
# KEY cleaning (STRONG) + COALESCE (CASE-INSENSITIVE)
# -----------------------------
SCI_RE = re.compile(r"^\s*-?\d+(\.\d+)?[eE][+-]?\d+\s*$")

def clean_key_series(s: pd.Series, treat_as_text: bool) -> pd.Series:
    """
    STRONG key cleaner for reconciliation keys.

    - Trims, removes commas/spaces (including hidden unicode spaces)
    - Removes Excel prefix "="
    - Strips quotes ' "
    - Fixes trailing .0 / .00
    - Converts scientific notation to full integer string
    - treat_as_text=True  -> keep ONLY A-Z and 0-9 (removes -,/,_,:,etc.)
    - treat_as_text=False -> keep only digits
    """

    def _one(v):
        if pd.isna(v):
            return ""
        t = str(v).strip()
        if t == "":
            return ""

        # remove hidden spaces
        t = t.replace("\u00A0", "").replace("\u200B", "").strip()

        # remove Excel "force text" prefix like ="12345"
        if t.startswith("="):
            t = t[1:].strip()

        # strip surrounding quotes
        t = t.strip('"').strip("'").strip()

        # remove commas/spaces/tabs/newlines
        t = re.sub(r"[\s,]+", "", t)

        # treat 0 / 0.0 / 0.00 as blank key
        if t in {"0", "0.0", "0.00", "0.000"}:
            return ""

        # fix excel float-like ids: 12345.0 / 12345.00
        t = re.sub(r"\.0+$", "", t)

        # scientific notation handling (Excel export issue)
        if SCI_RE.match(t):
            try:
                d = Decimal(t)
                t = format(d.quantize(Decimal(1)), "f")
            except (InvalidOperation, ValueError):
                pass

        if treat_as_text:
            # ✅ strongest normalization: keep only alphanumerics
            t = re.sub(r"[^0-9A-Za-z]+", "", t).upper()
            return t
        else:
            # numeric-only ids
            return re.sub(r"\D+", "", t)

    return s.apply(_one)

# -----------------------------
# ✅ COALESCE KEY (case-insensitive + flexible)
# -----------------------------
KEY_CANDIDATES_PRIORITY = [
    "lender loan account number",
    "loan account number",
    "loanaccountnumber",
    "lenderloanaccountnumber",
    "parent loan account number",
    "loan acct number",
    "loan acc number",
    "account number",
    "customer id",
    "loan id",
    "lender account number",
    "Transaction Id"
    "Transaction id"
    "TransactionId"
    "TransactionID"
]

# Smart key suggestion (scans column names, returns best-guess key cols)
_KEY_HINTS = [
    "lenderloanaccountnumber", "loanaccountnumber", "loanacctnumber",
    "loanaccnumber", "lan", "accountnumber", "loanid", "lenderaccountnumber",
    "customerid", "borrowerid", "accountid","Transaction Id"
]

def suggest_key_candidates(cols: list, top_k: int = 5) -> list:
    """
    Score column names and return best-guess key columns.
    Works for both standard names (loan account number) 
    and custom names (Reco, ID, Key, Reference etc.)
    """
    scored = []
    for c in cols:
        n = re.sub(r"[^a-z0-9]+", "", str(c).strip().lower())
        score = 0

        # ✅ Standard finance key hints — high score
        for h in _KEY_HINTS:
            if h in n:
                score += 10

        # ✅ Generic key-like endings
        if n.endswith("id") or n.endswith("number") or n.endswith("no"):
            score += 3
        if n.endswith("key") or n.endswith("code") or n.endswith("ref"):
            score += 3

        # ✅ Generic key-like prefixes
        if n.startswith("key") or n.startswith("id") or n.startswith("ref"):
            score += 3
        if n.startswith("reco") or n.startswith("unique") or n.startswith("primary"):
            score += 5

        # ✅ Short column names that look like keys (Reco, ID, Key, Ref)
        if len(n) <= 6 and any(x in n for x in ["reco", "key", "id", "ref", "lan", "acc"]):
            score += 5

        # ✅ Columns with numbers at end suggest versioning (Reco 1, ID2)
        # Give them lower priority than exact match but still include
        if re.match(r"^[a-z]+\d+$", n):
            score += 1

        # ✅ Penalize obvious non-key columns
        if any(x in n for x in [
            "date", "amount", "sum", "balance", "rate",
            "bonus", "target", "bcf", "collection",
            "status", "tenure", "roi", "principal",
            "outstanding", "pos", "dpd", "closure",
            "disbursal", "month", "year", "flag"
        ]):
            score -= 5

        scored.append((score, c))

    scored.sort(key=lambda x: (-x[0], str(x[1]).lower()))

    # ✅ Return top_k with score >= 1 (lowered from 2 to catch custom names)
    return [c for sc, c in scored if sc >= 1][:top_k]

def _norm_col_name(x: str):
    x = str(x).replace("\u00A0", " ").replace("\u200B", "")
    x = x.replace("\n", " ").replace("\r", " ")
    x = x.strip().lower()
    x = re.sub(r"\s+", " ", x)
    x_no_space = x.replace(" ", "")
    return x, x_no_space

def build_effective_key(df: pd.DataFrame, candidates: list):
    """
    Returns:
      raw_key_series: raw coalesced key (string)
      src_col_series: which column provided the key for that row

    ✅ Matches columns case-insensitively and also works if spaces/underscores differ.
    """
    if df is None or df.empty:
        return pd.Series([], dtype=str), pd.Series([], dtype=str)

    # map normalized names -> actual column name
    norm_map = {}
    for col in df.columns:
        a, b = _norm_col_name(col)
        norm_map[a] = col
        norm_map[b] = col

    raw_key = pd.Series([""] * len(df), index=df.index, dtype=str)
    src_col = pd.Series([""] * len(df), index=df.index, dtype=str)

    for cand in candidates:
        ca, cb = _norm_col_name(cand)
        actual = norm_map.get(ca) or norm_map.get(cb)
        if not actual or actual not in df.columns:
            continue

        s = df[actual]
        valid = ~s.apply(is_blank)
        fill_mask = (raw_key == "") & valid
        if fill_mask.any():
            raw_key.loc[fill_mask] = s.loc[fill_mask].astype(str)
            src_col.loc[fill_mask] = actual

    return raw_key, src_col

# -----------------------------
# Aggregation (OPTIMIZED)
# -----------------------------
def _first_non_blank(x: pd.Series):
    if x is None or len(x) == 0:
        return None
    for v in x:
        if not is_blank(v):
            return v
    return x.iloc[0]

def _join_unique_non_blank(x: pd.Series, sep="; "):
    vals = []
    seen = set()
    for v in x:
        if is_blank(v):
            continue
        sv = str(v)
        if sv not in seen:
            seen.add(sv)
            vals.append(sv)
    return sep.join(vals)

def aggregate_by_key_fast(df: pd.DataFrame, key_col: str, numeric_cols: list, join_cols: list) -> pd.DataFrame:
    """
    Aggregates only required columns.
    - numeric_cols: summed
    - join_cols: joined unique (for file/sheet trace)
    - others: first non-blank
    Adds _DUP_COUNT
    """
    df = df.copy()
    if key_col not in df.columns:
        return df

    dup_counts = df.groupby(key_col, dropna=False).size().rename("_DUP_COUNT").reset_index()

    agg_map = {}
    for c in df.columns:
        if c == key_col:
            continue
        if c in numeric_cols:
            agg_map[c] = "sum"
        elif c in join_cols:
            agg_map[c] = _join_unique_non_blank
        else:
            agg_map[c] = _first_non_blank

    out = df.groupby(key_col, dropna=False).agg(agg_map).reset_index()
    out = out.merge(dup_counts, on=key_col, how="left")
    return out

# -----------------------------
# Excel builder
# -----------------------------
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL   = PatternFill("solid", fgColor="E7EEF8")
DIFF_FILL     = PatternFill("solid", fgColor="E2F0D9")
MATCH_FILL    = PatternFill("solid", fgColor="D9E1F2")

BOLD   = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")

def _safe_df(df: pd.DataFrame) -> pd.DataFrame:
    return df if df is not None else pd.DataFrame()

def _write_plain_df(ws, df: pd.DataFrame, sheet_title: str):
    df = _safe_df(df)
    ws.title = sheet_title[:31]
    cols = list(df.columns)

    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    ws.freeze_panes = "A2"
    for j, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(45, max(12, len(str(col)) + 2))

def _write_table_with_subtotal_and_formulas(ws, df: pd.DataFrame, sheet_title: str, subtotal_label="Subtotal"):
    df = _safe_df(df)
    ws.title = sheet_title[:31]
    cols = list(df.columns)
    col_index = {c: idx + 1 for idx, c in enumerate(cols)}  # 1-based

    diff_cols  = [c for c in cols if c.startswith("Diff_") and c.endswith("_f1-f2")]
    match_cols = [c for c in cols if c.startswith("Match_") and c.endswith("_f1=f2")]

    # Row 2 headers
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=col)
        cell.font = BOLD
        cell.alignment = CENTER
        if col in diff_cols:
            cell.fill = DIFF_FILL
        elif col in match_cols:
            cell.fill = MATCH_FILL
        else:
            cell.fill = HEADER_FILL

    # Row 3+ data
    for i, row in enumerate(df.itertuples(index=False), start=3):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if cols[j-1] in diff_cols:
                c.fill = DIFF_FILL
            elif cols[j-1] in match_cols:
                c.fill = MATCH_FILL

    last_row = max(3, 2 + len(df))

    # Row 1 subtotal
    ws.cell(row=1, column=1, value=subtotal_label).font = BOLD
    ws.cell(row=1, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=1, column=1).alignment = LEFT

    for j, col in enumerate(cols, start=1):
        c_letter = get_column_letter(j)
        cell = ws.cell(row=1, column=j)
        cell.fill = SUBTOTAL_FILL
        cell.font = BOLD
        cell.alignment = CENTER
        if j == 1:
            continue

        series = df[col] if (len(df) > 0 and col in df.columns) else pd.Series([], dtype="object")
        numeric_series = pd.to_numeric(series, errors="coerce")
        is_num = numeric_series.notna().any() and (not col.startswith("Match_")) and (not col.startswith("Diff_") or col in diff_cols)
        if is_num:
            cell.value = f"=SUBTOTAL(9,{c_letter}3:{c_letter}{last_row})"
        else:
            cell.value = ""
    # Diff formulas — only for NUMERIC diff columns (skip text columns)
    for diff_col in diff_cols:
        base = diff_col[len("Diff_"):-len("_f1-f2")]
        c1 = f"{base}_f1"
        c2 = f"{base}_f2"
        if c1 in col_index and c2 in col_index:
            diff_j = col_index[diff_col]
            j1 = col_index[c1]
            j2 = col_index[c2]
            L1 = get_column_letter(j1)
            L2 = get_column_letter(j2)

            # ✅ FIX: check if the base column is actually numeric in the dataframe
            # If the Diff column in df is all blank/None/empty string, skip formula (text column)
            if diff_col in df.columns:
                diff_series = df[diff_col]
                numeric_check = pd.to_numeric(diff_series, errors="coerce")
                if not numeric_check.notna().any():
                    # Text column — leave Diff cells blank, do NOT write subtraction formula
                    for r in range(3, last_row + 1):
                        ws.cell(row=r, column=diff_j).value = ""
                    continue

            for r in range(3, last_row + 1):
                ws.cell(row=r, column=diff_j).value = f"={L1}{r}-{L2}{r}"

    # Match formulas
    for mcol in match_cols:
        base = mcol[len("Match_"):-len("_f1=f2")]
        c1 = f"{base}_f1"
        c2 = f"{base}_f2"
        if c1 in col_index and c2 in col_index:
            mj = col_index[mcol]
            j1 = col_index[c1]
            j2 = col_index[c2]
            L1 = get_column_letter(j1)
            L2 = get_column_letter(j2)
            for r in range(3, last_row + 1):
                ws.cell(row=r, column=mj).value = f"={L1}{r}={L2}{r}"

    for j, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(45, max(12, len(str(col)) + 2))

    ws.freeze_panes = "A3"

def build_excel_bytes(sheets: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for name, df in sheets.items():
        ws = wb.create_sheet(str(name)[:31])

        # Keep your existing formatting logic
        if name in ("Matched", "Mismatched"):
            _write_table_with_subtotal_and_formulas(ws, df, name)
        else:
            _write_plain_df(ws, df, name)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_summary_pdf(pdf_path: Path, summary_df: pd.DataFrame, app_version: str, focus_col: str):
    pdf_path = Path(pdf_path)

    from reportlab.lib.pagesizes import landscape, A4
    _page = landscape(A4)   # ✅ Force landscape explicitly

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=_page,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm
    )
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("<b> • Reconciliation Summary Report</b>", styles["Title"])
    subtitle = Paragraph(
        f"Tool Version: {app_version}<br/>"
        f"Generated On: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}<br/>"
        f"Summary Particular: {focus_col}",
        styles["Normal"]
    )

    story.append(title)
    story.append(Spacer(1, 8))
    story.append(subtitle)
    story.append(Spacer(1, 14))

    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    
    # ✅ Define styles for wrapped text
    _style_normal = ParagraphStyle(
        "normal_wrap",
        fontSize=8,
        leading=11,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )
    _style_center = ParagraphStyle(
        "center_wrap",
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        wordWrap="CJK",
    )
    _style_right = ParagraphStyle(
        "right_wrap",
        fontSize=8,
        leading=11,
        alignment=TA_RIGHT,
        wordWrap="CJK",
    )
    _style_header = ParagraphStyle(
        "header_wrap",
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.white,
        fontName="Helvetica-Bold",
        wordWrap="CJK",
    )

    # Format numeric columns to 2 decimals
    df_fmt = summary_df.copy()
    for col in df_fmt.columns:
        if pd.api.types.is_numeric_dtype(df_fmt[col]):
            df_fmt[col] = df_fmt[col].apply(
                lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
            )

    # ✅ Build table data with Paragraph objects for text wrapping
    # Column order: SNO, Particular, Count, f1, f2, Diff, Remarks
    _cols = list(df_fmt.columns)
    _header_row = [Paragraph(str(c), _style_header) for c in _cols]

    _data_rows = []
    for _, row in df_fmt.iterrows():
        _row_cells = []
        for i, val in enumerate(row):
            _val_str = str(val) if val is not None else ""
            if i == 0:  # SNO
                _row_cells.append(Paragraph(_val_str, _style_center))
            elif i == 1:  # Particular
                _row_cells.append(Paragraph(_val_str, _style_normal))
            elif i == len(_cols) - 1:  # Remarks (last col)
                _row_cells.append(Paragraph(_val_str, _style_normal))
            elif i in [2, 3, 4, 5]:  # Count, f1, f2, Diff
                _row_cells.append(Paragraph(_val_str, _style_right))
            else:
                _row_cells.append(Paragraph(_val_str, _style_normal))
        _data_rows.append(_row_cells)

    table_data = [_header_row] + _data_rows
    
    # ✅ Landscape A4 width = 277mm, minus margins 24mm = 253mm usable
    tbl = Table(
        table_data,
        repeatRows=1,
        colWidths=[10*mm, 38*mm, 22*mm, 45*mm, 45*mm, 33*mm, 60*mm]
    )
    tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        # Base table
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        # ✅ WRAP TEXT in all cells — fixes Remarks column truncation
        ("WORDWRAP", (0, 0), (-1, -1), True),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfd8e3")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbff")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),

        # Alignment
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),   # header row
        ("ALIGN", (0, 1), (0, -1), "CENTER"),   # SNO col
        ("ALIGN", (1, 1), (1, -1), "LEFT"),     # Particular col
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),    # Count col
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),    # f1 + f2 cols
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),    # Diff col
        ("ALIGN", (6, 1), (6, -1), "LEFT"),     # ✅ Remarks LEFT so wrap looks natural

        # Bold total row
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eaf2ff")),
    ]))

    story.append(tbl)
    story.append(Spacer(1, 14))

    footer = Paragraph(
        "This report is generated from Reconciliation Suite and runs locally on the user machine.",
        styles["Italic"]
    )
    story.append(footer)

    doc.build(story)

# -------------------------------------------------------------
# Single CSV bytes helper (keep only one definition)
# -------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _df_to_csv_bytes(_sig: str, df: pd.DataFrame) -> bytes:
    dfx = df if df is not None else pd.DataFrame()
    return dfx.to_csv(index=False).encode("utf-8")

# -----------------------------
# UI Upload
# -----------------------------
col1, col2 = st.columns(2)
with col1:
    f1_list = st.file_uploader(
        "Upload File 1 (multiple allowed)",
        type=["csv", "xlsx", "xls", "xlsb"],
        key="file1",
        accept_multiple_files=True
    )
with col2:
    f2_list = st.file_uploader(
        "Upload File 2 (multiple allowed)",
        type=["csv", "xlsx", "xls", "xlsb"],
        key="file2",
        accept_multiple_files=True
    )

if not (f1_list and f2_list):
    st.info("Upload at least 1 file on both sides to start.")
    st.stop()


# -----------------------------
# Read ONCE (per upload set)
# -----------------------------
sig1 = tuple((f.name, len(f.getvalue())) for f in f1_list)
sig2 = tuple((f.name, len(f.getvalue())) for f in f2_list)
sig = (sig1, sig2)
# ✅ upload signature for caching (search index / exports)
st.session_state["upload_signature"] = str(sig)

if st.session_state.get("files_sig") != sig:
    try:
        df1, f1_summary = read_many(f1_list)
        df2, f2_summary = read_many(f2_list)
    except Exception as e:
        st.error(f"Error reading files: {e}")
        st.stop()
    st.session_state["files_sig"] = sig
    st.session_state["df1"] = df1
    st.session_state["df2"] = df2
    st.session_state["f1_summary"] = f1_summary
    st.session_state["f2_summary"] = f2_summary
    st.session_state["confirmed_sig"] = None
    # ✅ Reset run state when uploads change
    st.session_state["run_done"] = False
    # MAIN (no matched) cache reset
    st.session_state["excel_ready_main"] = False
    st.session_state["excel_bytes_main"] = None
    st.session_state["excel_sig_main"] = None
    # MATCHED-only cache reset
    st.session_state["excel_ready_matched"] = False
    st.session_state["excel_bytes_matched"] = None
    st.session_state["excel_sig_matched"] = None
    # ✅ PATCH-3: Reset RUN cache when uploads change (prevents stale cached results)
    st.session_state["run_cache_ready"] = False
    st.session_state["run_cache_sig"] = None
    for k in [
        "summary_df", "key_diag_df",
        "matched_out", "mismatched_out",
        "only_f1_out", "only_f2_out",
        "dup_rows_1_out", "dup_rows_2_out", "dup_both_out",
        "miss_cols_f1", "miss_cols_f2",
        # ✅ f1_summary / f2_summary NOT deleted here — they are set by read_many above
    ]:
        if k in st.session_state:
            del st.session_state[k]

    # ✅ FIX 1: Reset dup mode radio and comparison cache when new files uploaded
    # This fixes the issue where Back to Launcher + relaunch keeps stale dup state
    for k in [
        "dup_mode_radio",
        "summary_incl_dups",
        "summary_excl_dups",
        "summary_incl_label",
        "summary_excl_label",
        "_has_dups",
        "_exclude_dups",
        "dup_exclude_mode",
        "dup_union_count",
        "dup_union",
        "dup_keys_1",
        "dup_keys_2",
        "dup_f1_key_count",
        "dup_f2_key_count",
    ]:
        if k in st.session_state:
            del st.session_state[k]
    # (Optional but recommended) Reset CSV cache on upload change
    for k in [
        "csv_sig_last",
        "csv_ready_matched", "csv_bytes_matched",
        "csv_ready_mismatched", "csv_bytes_mismatched",
        "csv_ready_only_f1", "csv_bytes_only_f1",
        "csv_ready_only_f2", "csv_bytes_only_f2",
        "csv_ready_zip_all", "csv_bytes_zip_all",
    ]:
        if k in st.session_state:
            del st.session_state[k]

else:
    df1 = st.session_state["df1"]
    df2 = st.session_state["df2"]
    f1_summary = st.session_state.get(
        "f1_summary",
        pd.DataFrame(columns=["File", "Sheet", "Rows", "Size_KB", "Time_Sec"])
    )
    f2_summary = st.session_state.get(
        "f2_summary",
        pd.DataFrame(columns=["File", "Sheet", "Rows", "Size_KB", "Time_Sec"])
    )

# -----------------------------
# Uploaded Files Summary + Matched Cols + Missing Cols by file
# -----------------------------
# ✅ Normalize column headers (fix invisible spaces/newlines) before comparing
df1 = normalize_cols(df1)
df2 = normalize_cols(df2)
st.session_state["df1"] = df1
st.session_state["df2"] = df2
# ── FREE PLAN ROW LIMIT CHECK ──
# ── FREE PLAN ROW LIMIT CHECK ──
if st.session_state.get("user_plan","free") == "free":
    _total = len(df1) + len(df2)
    if _total > 200:
        st.markdown("""
        <div style="background:#fff3cd; border:1px solid #ffc107; border-radius:12px;
                    padding:1rem 1.5rem; margin:0.5rem 0;">
            <b>🆓 Free Plan — Demo Mode</b><br>
            Showing first <b>100 rows from each file</b> for preview.
            Your files have <b>{:,} rows</b> total.
            Full reconciliation available on paid plans.
        </div>
        """.format(_total), unsafe_allow_html=True)
        df1 = df1.head(100)
        df2 = df2.head(100)
        st.session_state["df1"] = df1
        st.session_state["df2"] = df2
common_cols = sorted(list(set(df1.columns).intersection(set(df2.columns))))
internal_cols = {"_SOURCE_FILE", "_KEY", "_merge", "_SHEET"}
matched_cols = [c for c in common_cols if c not in internal_cols]

def missing_matched_cols_by_file(combined_df: pd.DataFrame, matched_cols: list) -> pd.DataFrame:
    if combined_df is None or combined_df.empty or "_SOURCE_FILE" not in combined_df.columns:
        return pd.DataFrame(columns=["File", "Missing_Matched_Cols_Count", "Missing_Matched_Cols"])

    out = []
    for fname, g in combined_df.groupby("_SOURCE_FILE", dropna=False):
        miss = []
        for c in matched_cols:
            if c not in g.columns:
                miss.append(c)
                continue
            s = g[c]
            if s.isna().all():
                miss.append(c)

        out.append({
            "File": str(fname),
            "Missing_Matched_Cols_Count": int(len(miss)),
            "Missing_Matched_Cols": ", ".join(miss)
        })

    df_out = pd.DataFrame(out)
    if df_out.empty:
        return pd.DataFrame(columns=["File", "Missing_Matched_Cols_Count", "Missing_Matched_Cols"])
    return df_out.sort_values(["Missing_Matched_Cols_Count", "File"], ascending=[False, True])

miss_cols_f1 = missing_matched_cols_by_file(df1, matched_cols)
miss_cols_f2 = missing_matched_cols_by_file(df2, matched_cols)

st.subheader("0) Uploaded Files Summary")
m1, m2, m3 = st.columns(3)
m1.metric("File 1 - Total files", len(f1_list))
m2.metric("File 2 - Total files", len(f2_list))
m3.metric("Matched Columns", len(matched_cols))

with st.expander("Show matched column names"):
    st.write(matched_cols)

left, right = st.columns(2)
with left:
    st.write("File 1 uploads")
    st.caption(f"Total files: {len(f1_list)} | Total rows: {len(df1):,}")
    st.dataframe(f1_summary, use_container_width=True, hide_index=True)

with right:
    st.write("File 2 uploads")
    st.caption(f"Total files: {len(f2_list)} | Total rows: {len(df2):,}")
    st.dataframe(f2_summary, use_container_width=True, hide_index=True)

st.markdown("### 🔍 Missing Matched Columns (file-wise)")
cA, cB = st.columns(2)
with cA:
    st.write("File 1 side: files missing any of the matched columns")
    st.dataframe(miss_cols_f1, use_container_width=True, hide_index=True)
with cB:
    st.write("File 2 side: files missing any of the matched columns")
    st.dataframe(miss_cols_f2, use_container_width=True, hide_index=True)

# ============================================================
# ✅ 0.5) COLUMN MAPPING — map unmatched columns across files
# ============================================================

st.markdown("### 🔀 Column Mapping (optional)")

# Find unmatched columns
_unmatched_f1 = sorted([c for c in df1.columns if c not in internal_cols and c not in matched_cols])
_unmatched_f2 = sorted([c for c in df2.columns if c not in internal_cols and c not in matched_cols])

if not _unmatched_f1 and not _unmatched_f2:
    # All columns already match — no mapping needed
    st.success("✅ All columns match perfectly — no mapping needed!")
    st.session_state["col_mapping"] = {}
else:
    # Show checkbox — mapping UI only appears when ticked
    _mapping_enabled = st.checkbox(
        f"🔀 Enable Column Mapping  ({len(_unmatched_f1)} unmatched in File 1 | {len(_unmatched_f2)} unmatched in File 2)",
        value=False,
        key="mapping_enabled"
    )

    if not _mapping_enabled:
        # ✅ Checkbox OFF — hide everything, clear any old mapping
        st.caption("💡 Tick above if your files have the same data under different column names.")
        st.session_state["col_mapping"] = {}

    else:
        # ✅ Checkbox ON — show full mapping UI
        st.caption("Map File 1 columns → File 2 columns. Leave as '— Skip —' to ignore.")

        # Auto-suggest obvious matches by normalized name
        def _auto_suggest(cols_f1, cols_f2):
            def _n(c):
                return re.sub(r"[^a-z0-9]+", "", str(c).strip().lower())
            f2_norm = {_n(c): c for c in cols_f2}
            return {c1: f2_norm[_n(c1)] for c1 in cols_f1 if _n(c1) in f2_norm}

        _auto = _auto_suggest(_unmatched_f1, _unmatched_f2)

        # Reset mapping if files changed
        if st.session_state.get("col_mapping_sig") != str(sig):
            st.session_state["col_mapping_sig"] = str(sig)
            st.session_state["col_mapping"] = _auto.copy()

        if _auto:
            st.success(f"✅ Auto-suggested {len(_auto)} mapping(s) based on similar names.")

        st.info(
            f"📊 **{len(matched_cols)}** already matched | "
            f"**{len(_unmatched_f1)}** unmatched in File 1 | "
            f"**{len(_unmatched_f2)}** unmatched in File 2"
        )

        st.markdown("#### Map File 1 → File 2")
        _f2_options = ["— Skip —"] + _unmatched_f2
        _new_mapping = {}

        for i, f1_col in enumerate(_unmatched_f1):
            _ca, _cb, _cc = st.columns([2, 0.3, 2])
            with _ca:
                st.markdown(
                    f"<div style='padding:6px 10px;background:#eef3ff;border-radius:8px;"
                    f"font-size:13px;font-weight:600;color:#0b5ed7;border:1px solid #c8d8f8;'>"
                    f"📄 {f1_col}</div>",
                    unsafe_allow_html=True
                )
            with _cb:
                st.markdown("<div style='text-align:center;padding-top:6px;font-size:18px;'>→</div>", unsafe_allow_html=True)
            with _cc:
                _cur = st.session_state["col_mapping"].get(f1_col, _auto.get(f1_col, "— Skip —"))
                _idx = _f2_options.index(_cur) if _cur in _f2_options else 0
                _sel = st.selectbox(
                    "Map to",
                    options=_f2_options,
                    index=_idx,
                    key=f"colmap_{i}_{f1_col}",
                    label_visibility="collapsed"
                )
                if _sel != "— Skip —":
                    _new_mapping[f1_col] = _sel

            st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

        st.session_state["col_mapping"] = _new_mapping

        if _new_mapping:
            st.markdown("#### ✅ Active Mappings")
            st.dataframe(
                pd.DataFrame([{"File 1 Column": k, "→": "→", "File 2 Column": v} for k, v in _new_mapping.items()]),
                use_container_width=True, hide_index=True
            )
            st.caption(f"✅ {len(_new_mapping)} column(s) mapped — will be reconciled together.")
        else:
            st.info("No mappings set yet. Use the dropdowns above to map columns.")

# ============================================================
# ✅ Apply column mapping to df2 (always runs — empty mapping = no change)
# ============================================================
col_mapping = st.session_state.get("col_mapping", {})

if col_mapping:
    _rename_map = {v: k for k, v in col_mapping.items()}
    df2 = df2.rename(columns=_rename_map)
    st.session_state["df2"] = df2

    # Recompute matched cols after mapping
    common_cols = sorted(list(set(df1.columns).intersection(set(df2.columns))))
    matched_cols = [c for c in common_cols if c not in internal_cols]

    st.success(
        f"✅ Mapping applied! Matched columns: **{len(matched_cols)}** total "
        f"({len(matched_cols) - len([c for c in matched_cols if c not in _unmatched_f1])} original + {len(col_mapping)} mapped)"
    )





# -----------------------------
# Preview (combined)
# -----------------------------
st.subheader("1) Preview (combined)")
p1, p2 = st.columns(2)
with p1:
    st.write("File 1 Combined Preview")
    st.dataframe(df1.head(20), use_container_width=True, hide_index=True)
with p2:
    st.write("File 2 Combined Preview")
    st.dataframe(df2.head(20), use_container_width=True, hide_index=True)

if not common_cols:
    st.error("No common columns found between the two sides. Please ensure headers match.")
    st.stop()

# -----------------------------
# ✅ COALESCE KEY MODE — Smart user-selectable key (from app_v2 logic)
# -----------------------------
def _norm_name(x: str) -> str:
    x = str(x).replace("\u00A0", " ").replace("\u200B", "")
    x = x.replace("\n", " ").replace("\r", " ")
    x = x.strip().lower()
    x = re.sub(r"\s+", " ", x)
    x_no_space = x.replace(" ", "")
    return x_no_space

common_map = { _norm_name(c): c for c in common_cols }

# ✅ Step 1: check which priority candidates exist in both files
def _dedup(lst):
    seen = set(); out = []
    for x in lst:
        if x not in seen: seen.add(x); out.append(x)
    return out

existing_key_candidates = []
for cand in KEY_CANDIDATES_PRIORITY:
    k = _norm_name(cand)
    if k in common_map:
        existing_key_candidates.append(common_map[k])
# ✅ Dedup immediately — KEY_CANDIDATES_PRIORITY may map to same actual column
existing_key_candidates = _dedup(existing_key_candidates)

# ✅ Step 2: if none found, smart-suggest from column names
# This now works for custom column names like Reco, Reco 1, ID, Reference etc.
auto_suggest = existing_key_candidates if existing_key_candidates else _dedup(suggest_key_candidates(common_cols, top_k=3))
st.markdown("### 🔑 Key Column (used for matching rows)")

if existing_key_candidates:
    st.success("✅ Default key column(s) found → " + " / ".join(existing_key_candidates))
else:
    # ✅ Show smart suggestions even when default names not found
    _smart_suggest = suggest_key_candidates(common_cols, top_k=5)
    if _smart_suggest:
        st.warning(
            f"⚠️ Default key names (like 'Lender Loan Account Number') not found in your files. \n\n"
            f"**Smart suggestion based on your column names:** {' / '.join(_smart_suggest)} \n\n"
            f"Please select the correct key column(s) from the multiselect below."
        )
    else:
        st.warning(
            "⚠️ Could not auto-detect key columns. "
            "Please manually select the correct key column(s) from the list below."
        )
with st.expander("ℹ️ How COALESCE key works"):
    st.write(
        "The tool builds a single _KEY as the first non-blank value from your selected "
        "key columns (in the order shown). This handles files where one column may be blank."
    )

# ✅ Step 3: User can select / override key columns from ALL common columns
# Deduplicate auto_suggest so same column doesn't appear twice in default
auto_suggest = _dedup(auto_suggest)

# ✅ Deduplicate default — ensure same column name never appears twice
_safe_default = []
_seen_default = set()
for c in auto_suggest:
    if c in common_cols and c not in _seen_default:
        _seen_default.add(c)
        _safe_default.append(c)

key_cols_sel = st.multiselect(
    "Select key column(s) — COALESCE priority follows order below",
    options=_dedup(common_cols),
    default=_safe_default
)

priority_text = st.text_input(
    "Key priority order (comma-separated, left = highest priority). Edit if needed.",
    value=", ".join(_dedup(key_cols_sel)) if key_cols_sel else ", ".join(auto_suggest)
)

priority_list = _dedup([x.strip() for x in priority_text.split(",") if x.strip()])
_allowed_keys = set(key_cols_sel) if key_cols_sel else set(existing_key_candidates)
existing_key_candidates = _dedup([c for c in priority_list if c in common_cols and c in _allowed_keys])
if not existing_key_candidates:
    existing_key_candidates = _dedup([c for c in (key_cols_sel or auto_suggest) if c in common_cols])

if not existing_key_candidates:
    st.warning("⚠️ No key selected yet. Please select at least one key column above.")
    st.stop()

st.info("✅ Key Mode: COALESCE → " + " / ".join(existing_key_candidates))

# -----------------------------
# 🔎 Key Validation (Detected Key Column(s) + Raw vs Clean Preview)
# -----------------------------
st.markdown("### 🔎 Key Validation")

# Which key columns exist on each side (sanity)
det_f1 = [c for c in existing_key_candidates if c in df1.columns]
det_f2 = [c for c in existing_key_candidates if c in df2.columns]

k1, k2 = st.columns(2)
with k1:
    st.write("**File 1 — Detected Key Column(s)**")
    st.write(det_f1 if det_f1 else "❌ None detected")
with k2:
    st.write("**File 2 — Detected Key Column(s)**")
    st.write(det_f2 if det_f2 else "❌ None detected")

with st.expander("Show Raw vs Clean Key Preview (Top 20)", expanded=False):
    # Use only a small slice to keep it fast on big files
    _df1_small = df1.head(500).copy()
    _df2_small = df2.head(500).copy()

    raw1, src1 = build_effective_key(_df1_small, existing_key_candidates)
    raw2, src2 = build_effective_key(_df2_small, existing_key_candidates)

    # Treat as text for loan account keys (alphanumeric safe)
    clean1 = clean_key_series(raw1, treat_as_text=True)
    clean2 = clean_key_series(raw2, treat_as_text=True)

    prev1 = pd.DataFrame({
        "Key_Source_Col": src1,
        "Raw_Key": raw1,
        "Clean_Key": clean1,
    })
    prev2 = pd.DataFrame({
        "Key_Source_Col": src2,
        "Raw_Key": raw2,
        "Clean_Key": clean2,
    })

    # show top 20 non-blank keys
    prev1_nz = prev1[prev1["Clean_Key"] != ""].head(20)
    prev2_nz = prev2[prev2["Clean_Key"] != ""].head(20)

    c_prev1, c_prev2 = st.columns(2)
    with c_prev1:
        st.write("**File 1 — Top 20**")
        st.dataframe(prev1_nz, use_container_width=True, hide_index=True)
    with c_prev2:
        st.write("**File 2 — Top 20**")
        st.dataframe(prev2_nz, use_container_width=True, hide_index=True)

    # quick overlap signal
    set1 = set(prev1_nz["Clean_Key"].tolist())
    set2 = set(prev2_nz["Clean_Key"].tolist())
    overlap = len(set1.intersection(set2))
    st.caption(f"Preview overlap (Top 20 cleaned keys): **{overlap}** common keys")


# -----------------------------
# 🔍 Search Loan Account Number (Raw → Clean + exists in both sides)
# -----------------------------
st.markdown("### 🔍 Search Loan Account Number")

search_lan = st.text_input(
    "Type any Loan Account Number / Lender Loan Account Number",
    placeholder="e.g., PYTMML_MFL1242410533728 or MFLMFOPAYTL000005333725"
)

# Clean the input the same way we clean reconciliation keys
if search_lan:
    _in_series = pd.Series([search_lan])
    _clean_in = clean_key_series(_in_series, treat_as_text=True).iloc[0]
    st.write(f"**Raw:** `{search_lan}`")
    st.write(f"**Cleaned:** `{_clean_in}`")

    def build_clean_key_set(df: pd.DataFrame, candidates: list, chunk_size: int = 200000) -> set:
        """
        Build a set of CLEANED coalesced keys for quick membership checks.
        """
        raw_key, _ = build_effective_key(df, candidates)
        n = len(raw_key)
        out = set()
        if n == 0:
            return out
        prog = st.progress(0, text="Building key index for search...")
        for start in range(0, n, chunk_size):
            end = min(start + chunk_size, n)
            chunk = raw_key.iloc[start:end]
            cleaned = clean_key_series(chunk, treat_as_text=True)
            out.update([k for k in cleaned.tolist() if k])
            prog.progress(int(end * 100 / n), text=f"Building key index for search... {end:,}/{n:,}")
        prog.empty()
        return out

    # ── Cache key: upload sig + candidates
    _sig = st.session_state.get("upload_signature", "")
    _cand_sig = "|".join(existing_key_candidates)

    index_ready = False
    if "search_key_index_meta" in st.session_state:
        meta = st.session_state["search_key_index_meta"]
        if meta.get("sig") == _sig and meta.get("cand_sig") == _cand_sig:
            index_ready = True

    # ── AUTO-BUILD: trigger silently when user types a key (no button needed)
    if not index_ready:
        with st.spinner("Building search index (one-time for this upload)…"):
            st.session_state["search_key_set_1"] = build_clean_key_set(df1, existing_key_candidates)
            st.session_state["search_key_set_2"] = build_clean_key_set(df2, existing_key_candidates)
            st.session_state["search_key_index_meta"] = {"sig": _sig, "cand_sig": _cand_sig}
            index_ready = True

    # ── LOOKUP — always show result immediately after typing
    s1 = st.session_state.get("search_key_set_1", set())
    s2 = st.session_state.get("search_key_set_2", set())

    in1 = _clean_in in s1
    in2 = _clean_in in s2

    cA, cB, cC = st.columns(3)
    with cA:
        st.metric("Exists in File 1", "✅ Yes" if in1 else "❌ No")
    with cB:
        st.metric("Exists in File 2", "✅ Yes" if in2 else "❌ No")
    with cC:
        if in1 and in2:
            st.success("✅ Present in BOTH files (after cleaning).")
        elif in1 and not in2:
            st.warning("⚠️ Present only in File 1 (after cleaning).")
        elif (not in1) and in2:
            st.warning("⚠️ Present only in File 2 (after cleaning).")
        else:
            st.error("❌ Not found in either file (after cleaning).")


# Keep a display label (not used alone for key build)
best_key_col = existing_key_candidates[0]

# -----------------------------
# Settings
# -----------------------------
def default_compare_selection(common_cols):
    desired_order = [
        "Lender Loan Account Number",
        "loan account number",
        "Loan Account Number",
        "CPB",
        "Target",
        "Target Rate",
        "Total BCF",
        "Total Collection",
        "Total Target",
        "loan amount",
        "Loan amount",
        "ROI",
        "TENURE",
        "Principal Outstanding",
        "Principal_ Outstanding",
    ]
    s = set(common_cols)
    out = []
    for d in desired_order:
        if d in s and d not in out:
            out.append(d)
    return out

def reorder_selected_cols(compare_cols):
    priority = [
        "Lender Loan Account Number",
        "loan account number",
        "Loan Account Number",
        "Lender account number",
        "lender account number",
    ]
    pri = [c for c in priority if c in compare_cols]
    rest = [c for c in compare_cols if c not in pri]
    return pri + rest

st.subheader("2) Select Reco Settings")
c1, c2, c3 = st.columns([2, 2, 2])

with c2:
    compare_mode = st.selectbox("Compare mode", ["Compare selected columns", "Compare common columns"])
    default_sel = default_compare_selection(common_cols)

    if compare_mode == "Compare selected columns":
        compare_cols = st.multiselect(
            "Select columns to compare (must exist in both)",
            common_cols,
            default=default_sel if default_sel else (common_cols[:7] if len(common_cols) >= 7 else common_cols)
        )
        compare_cols = reorder_selected_cols(compare_cols)
        reco_focus_col = st.selectbox(
            "Summary 'Particular' column (choose ONE from selected columns)",
            options=compare_cols if compare_cols else common_cols,
            index=0 if compare_cols else 0
        )
        # ✅ Move selected Particular column to FRONT of compare list (first in output columns)
        if reco_focus_col and reco_focus_col in compare_cols:
            rest = [c for c in compare_cols if c != reco_focus_col]
            compare_cols = [reco_focus_col] + rest
    else:
        compare_cols = common_cols
        reco_focus_col = st.selectbox(
            "Summary 'Particular' column (choose ONE from common columns)",
            options=common_cols,
            index=0
        )

with c3:
    numeric_tolerance = st.number_input(
        "Numeric tolerance (+/-) default 5",
        min_value=0.0,
        value=5.0,
        step=1.0
    )
    treat_blanks_as_equal = st.checkbox("Treat blanks/NaN as equal", value=True)

    st.markdown("**Performance options**")
    skip_mismatch_columns = st.checkbox("✅ Skip Mismatch_Columns (fastest mode)", value=True)
    show_rows_in_browser = st.number_input("Rows to show in browser (per tab)", min_value=100, value=2000, step=500)

with c1:
    st.write("✅ Key is COALESCE (no detection).")
    st.caption("If Lender Loan Account Number is blank, uses loan account number automatically.")
    st.caption("Only-in File1/File2 are never counted as mismatches.")
    st.caption("Step-2 aggregation is optimized to only required columns.")

if compare_mode == "Compare selected columns" and not compare_cols:
    st.warning("Please select at least one column to compare.")
    st.stop()

# ============================================================
# 🧹 Optional Filter — run reco on a SUBSET of rows
# ============================================================
st.markdown("### 🧹 Optional Filter (run reco on subset)")
st.caption("Example: if you have a Lender ID column in both files, filter to run reco only for one lender.")

filter_enabled = st.checkbox("Enable filter (applied on BOTH sides before reco)", value=False)

# Smart default: try to find a lender-like column
_filter_candidates = [
    "lender_id", "lender", "lender code", "lender_code",
    "lender name", "lender_name", "partner_id", "partner",
    "merchant_id", "merchant"
]
_filter_default = None
_common_lower = {c.strip().lower(): c for c in common_cols}
for _cand in _filter_candidates:
    if _cand.lower() in _common_lower:
        _filter_default = _common_lower[_cand.lower()]
        break
if _filter_default is None and common_cols:
    _filter_default = common_cols[0]

_filter_col_idx = common_cols.index(_filter_default) if _filter_default in common_cols else 0
filter_col = st.selectbox("Filter column", options=common_cols, index=_filter_col_idx)

# Show unique value picker + manual text entry
picked_values = []
if filter_enabled:
    use_picker = st.checkbox("Show available values (dropdown)", value=True)
    if use_picker:
        @st.cache_data(show_spinner=False)
        def _get_unique_filter_vals(_sig, col, limit=5000):
            """Get unique values from df1+df2 for the filter column (cached by upload sig)."""
            vals = []
            for _df in [df1, df2]:
                if col in _df.columns:
                    for v in _df[col].dropna().astype(str).str.strip().unique():
                        if v and v not in vals:
                            vals.append(v)
                        if len(vals) >= limit:
                            return vals, True
            vals.sort()
            return vals, False

        _fvals, _truncated = _get_unique_filter_vals(
            st.session_state.get("files_sig", ""), filter_col
        )
        if _truncated:
            st.info("Showing first 5,000 unique values. You can still type values manually below.")
        if not _fvals:
            st.info("No values found for this column.")
        picked_values = st.multiselect(
            "Pick filter value(s) (case-insensitive match)",
            options=_fvals,
            default=[]
        )

filter_value_raw = st.text_area(
    "Filter value(s) — one per line or comma-separated (case-insensitive, exact match after trim)",
    value="",
    height=90,
    placeholder="Example:\n85\n86\n87   (or  85, 86, 87)"
)

_tmp_vals = []
for _part in filter_value_raw.replace(",", "\n").splitlines():
    _v = _part.strip()
    if _v and _v not in _tmp_vals:
        _tmp_vals.append(_v)
filter_values = _tmp_vals
for _v in (picked_values or []):
    _vv = str(_v).strip()
    if _vv and _vv not in filter_values:
        filter_values.append(_vv)

if filter_enabled and not filter_values:
    st.warning("⚠️ Filter is enabled but no value entered. Please add at least one value or disable the filter.")
elif filter_enabled and filter_values:
    st.caption(f"✅ Filter will apply on **{len(filter_values)}** value(s) → column: **{filter_col}**")

# -----------------------------
# ✅ 2.5) Selected Field Summary (before Run) — FIXED TO MATCH RUN LOGIC + COALESCE KEY
# -----------------------------
def build_selected_field_summary_before_run(df_side: pd.DataFrame, focus_col: str) -> pd.DataFrame:
    """
    IMPORTANT: This summary uses the SAME key-building as the Run:
    - COALESCE key from multiple columns
    - clean key with alphanumerics preserved
    - drops blank/invalid keys
    """
    if df_side is None or df_side.empty:
        return pd.DataFrame(columns=["File", "Sheet", "Rows", f"{focus_col}_Sum"])

    if "_SOURCE_FILE" not in df_side.columns:
        return pd.DataFrame(columns=["File", "Sheet", "Rows", f"{focus_col}_Sum"])

    tmp = df_side.copy()
    if "_SHEET" not in tmp.columns:
        tmp["_SHEET"] = "UNKNOWN"

    # ✅ Use user-selected keys not hardcoded list
    _key_priority = existing_key_candidates if existing_key_candidates else KEY_CANDIDATES_PRIORITY
    raw_key, src_col = build_effective_key(tmp, _key_priority)
    tmp["_KEY_TMP_SOURCE_COL"] = src_col
    tmp["_KEY_TMP"] = clean_key_series(raw_key, treat_as_text=True)

    tmp = tmp[tmp["_KEY_TMP"] != ""].copy()

    # focus might not exist in some files; coerce missing to 0
    if focus_col not in tmp.columns:
        tmp[focus_col] = 0

    tmp["_FOCUS_NUM"] = pd.to_numeric(tmp[focus_col], errors="coerce").fillna(0)

    g = tmp.groupby(["_SOURCE_FILE", "_SHEET"], dropna=False).agg(
        Rows=("_FOCUS_NUM", "size"),
        Focus_Sum=("_FOCUS_NUM", "sum")
    ).reset_index()

    out = g.rename(columns={
        "_SOURCE_FILE": "File",
        "_SHEET": "Sheet",
        "Focus_Sum": f"{focus_col}_Sum"
    }).sort_values(["Rows", "File"], ascending=[False, True])

    grand = {
        "File": "GRAND TOTAL",
        "Sheet": "",
        "Rows": int(out["Rows"].sum()),
        f"{focus_col}_Sum": float(out[f"{focus_col}_Sum"].sum())
    }
    out = pd.concat([out, pd.DataFrame([grand])], ignore_index=True)
    return out

st.subheader("2.5) Selected Field Summary (before Run)")

sf_left, sf_right = st.columns(2)
with sf_left:
    st.caption(f"File 1 side: {reco_focus_col} file-wise sum (valid keys only)")
    f1_focus_summary = build_selected_field_summary_before_run(df1, reco_focus_col)
    st.dataframe(f1_focus_summary, use_container_width=True, hide_index=True)

with sf_right:
    st.caption(f"File 2 side: {reco_focus_col} file-wise sum (valid keys only)")
    f2_focus_summary = build_selected_field_summary_before_run(df2, reco_focus_col)
    st.dataframe(f2_focus_summary, use_container_width=True, hide_index=True)

# -----------------------------
# Run
# -----------------------------
st.subheader("3) Run")

if "run_done" not in st.session_state:
    st.session_state["run_done"] = False

if st.button("✅ Run Reco", type="primary"):
    st.session_state["run_done"] = True

if not st.session_state["run_done"]:
    st.stop()

# -----------------------------
# ✅ Prevent "loop" on downloads: cache RUN outputs in session_state
# Any click (download/button) reruns Streamlit. So we reuse computed outputs.
# -----------------------------
_filter_sig = (
    str(filter_enabled) + "|" +
    (filter_col if filter_enabled else "") + "|" +
    ",".join(sorted([str(v).strip().upper() for v in filter_values])) if filter_enabled else "nofilter"
)

# ✅ Include dup_mode in signature so switching Include/Exclude triggers fresh run
_dup_mode_sig = str(st.session_state.get("dup_mode_radio", "Include"))

run_inputs_sig = (
    str(st.session_state.get("files_sig")) + "|" +
    str(reco_focus_col) + "|" +
    str(compare_mode) + "|" +
    ",".join(list(compare_cols)) + "|" +
    str(float(numeric_tolerance)) + "|" +
    str(bool(treat_blanks_as_equal)) + "|" +
    str(bool(skip_mismatch_columns)) + "|" +
    _filter_sig + "|" +
    _dup_mode_sig
)

RUN_CACHE_KEYS = [
    "summary_df", "key_diag_df",
    "matched_out", "mismatched_out",
    "only_f1_out", "only_f2_out",
    "dup_rows_1_out", "dup_rows_2_out", "dup_both_out",
    "miss_cols_f1", "miss_cols_f2",
    "f1_summary", "f2_summary",
    "dup_union", "dup_keys_1", "dup_keys_2",
    "dup_f1_key_count", "dup_f2_key_count",
    # ✅ ADD: track which mode was used so UI shows correct tip
    "_has_dups", "_exclude_dups"
]
has_cached_run = (
    st.session_state.get("run_cache_sig") == run_inputs_sig and
    st.session_state.get("run_cache_ready") is True and
    all(k in st.session_state for k in RUN_CACHE_KEYS)
)

skip_recompute = has_cached_run  # ✅ the missing piece in your file earlier

# Progress UI (used in both cached & fresh runs)
prog = st.progress(0)
label = st.empty()

# ✅ Run log — captures each step message for the Run Log tab
if "run_log_lines" not in st.session_state:
    st.session_state["run_log_lines"] = []

def progress_update(step, total_steps, msg):
    from datetime import datetime as _dt
    pct = int((step / total_steps) * 100)
    prog.progress(min(100, pct))
    label.info(f"Step {step}/{total_steps}: {msg}")
    ts = _dt.now().strftime("%H:%M:%S")
    st.session_state["run_log_lines"].append(f"[{ts}] Step {step}/{total_steps}: {msg}")

TOTAL_STEPS = 7


if has_cached_run:
    # ✅ Reuse outputs (NO recomputation on download click)
    summary_df      = st.session_state["summary_df"]
    key_diag_df     = st.session_state["key_diag_df"]
    matched_out     = st.session_state["matched_out"]
    mismatched_out  = st.session_state["mismatched_out"]
    only_f1_out     = st.session_state["only_f1_out"]
    only_f2_out     = st.session_state["only_f2_out"]
    dup_rows_1_out  = st.session_state["dup_rows_1_out"]
    dup_rows_2_out  = st.session_state["dup_rows_2_out"]
    dup_both_out    = st.session_state["dup_both_out"]
    miss_cols_f1    = st.session_state["miss_cols_f1"]
    miss_cols_f2    = st.session_state["miss_cols_f2"]
    f1_summary      = st.session_state["f1_summary"]
    f2_summary      = st.session_state["f2_summary"]
    # ✅ FIX: restore dup_union + dup_keys_1 + dup_keys_2 from cache so downloads work
    dup_union        = st.session_state.get("dup_union", set())
    dup_keys_1       = st.session_state.get("dup_keys_1", set())
    dup_keys_2       = st.session_state.get("dup_keys_2", set())
    dup_f1_key_count = st.session_state.get("dup_f1_key_count", len(dup_keys_1))
    dup_f2_key_count = st.session_state.get("dup_f2_key_count", len(dup_keys_2))
    _has_dups        = st.session_state.get("_has_dups", False)
    _exclude_dups    = st.session_state.get("_exclude_dups", False)
    st.info("✅ Using cached Run results (no recomputation on download).")
    progress_update(6, TOTAL_STEPS, "Using cached results (skipping recompute)...")

# ============================================================
# Steps 1–6: run ONLY when cache is NOT available
# ============================================================
if not skip_recompute:
    # -----------------------------
    # Step 1/7: RAW keys + Duplicate sheets (COALESCE KEY)
    # -----------------------------
    progress_update(1, TOTAL_STEPS, "Building RAW keys + extracting duplicates...")

    df1_raw = df1.copy()
    df2_raw = df2.copy()

    # ✅ Apply optional filter to BOTH sides before reco
    if filter_enabled and filter_values and filter_col:
        _fvals_upper = [str(v).strip().upper() for v in filter_values if str(v).strip()]
        if _fvals_upper:
            # resolve actual column name case-insensitively
            _f1_col = next((c for c in df1_raw.columns if c.strip().lower() == filter_col.strip().lower()), None)
            _f2_col = next((c for c in df2_raw.columns if c.strip().lower() == filter_col.strip().lower()), None)
            if _f1_col:
                df1_raw = df1_raw[df1_raw[_f1_col].astype(str).str.strip().str.upper().isin(_fvals_upper)].copy()
            if _f2_col:
                df2_raw = df2_raw[df2_raw[_f2_col].astype(str).str.strip().str.upper().isin(_fvals_upper)].copy()
            st.info(f"🔍 Filter applied: **{filter_col}** IN {_fvals_upper[:5]}{'...' if len(_fvals_upper)>5 else ''} → File1: {len(df1_raw):,} rows | File2: {len(df2_raw):,} rows")
            if df1_raw.empty or df2_raw.empty:
                st.error("⚠️ Filter returned 0 rows on one or both sides. Please check your filter value.")
                st.stop()

    # ✅ FIX: Use SELECTED key columns (existing_key_candidates) not hardcoded list
    # This means if user selects "Reco" and "Reco 1" those will be used as keys
    # existing_key_candidates is already set from user selection above
    _runtime_key_priority = existing_key_candidates if existing_key_candidates else KEY_CANDIDATES_PRIORITY
    raw1, src1 = build_effective_key(df1_raw, _runtime_key_priority)
    raw2, src2 = build_effective_key(df2_raw, _runtime_key_priority)
    df1_raw["_KEY_SOURCE_COL"] = src1
    df2_raw["_KEY_SOURCE_COL"] = src2

    # Loan account keys are TEXT-like → preserve alphanumerics
    df1_raw["_KEY"] = clean_key_series(raw1, treat_as_text=True)
    df2_raw["_KEY"] = clean_key_series(raw2, treat_as_text=True)

    df1_raw = df1_raw[df1_raw["_KEY"] != ""].copy()
    df2_raw = df2_raw[df2_raw["_KEY"] != ""].copy()

    c1_counts = df1_raw.groupby("_KEY").size()
    c2_counts = df2_raw.groupby("_KEY").size()

    dup_keys_1 = set(c1_counts[c1_counts > 1].index.tolist())
    dup_keys_2 = set(c2_counts[c2_counts > 1].index.tolist())

    dup_rows_1 = df1_raw[df1_raw["_KEY"].isin(dup_keys_1)].copy()
    dup_rows_2 = df2_raw[df2_raw["_KEY"].isin(dup_keys_2)].copy()

    dup_rows_1.insert(0, "Source", "File1")
    dup_rows_2.insert(0, "Source", "File2")

    dup_rows_1 = dup_rows_1.sort_values(["_KEY"])
    dup_rows_2 = dup_rows_2.sort_values(["_KEY"])

    keys1_all = set(df1_raw["_KEY"].unique())
    keys2_all = set(df2_raw["_KEY"].unique())
    keys_in_both = keys1_all.intersection(keys2_all)

    dup_union = dup_keys_1.union(dup_keys_2)
    dup_both_keys = dup_union.intersection(keys_in_both)

    # ✅ FIX: store exact unique key counts for summary + downloads
    dup_f1_key_count = len(dup_keys_1)   # unique keys duplicated in File1
    dup_f2_key_count = len(dup_keys_2)   # unique keys duplicated in File2

    dup_both_1 = df1_raw[df1_raw["_KEY"].isin(dup_both_keys)].copy()
    dup_both_2 = df2_raw[df2_raw["_KEY"].isin(dup_both_keys)].copy()
    dup_both_1.insert(0, "Source", "File1")
    dup_both_2.insert(0, "Source", "File2")
    dup_both = pd.concat([dup_both_1, dup_both_2], ignore_index=True).sort_values(["_KEY", "Source"])

    # show which candidate columns exist (for duplicate tab columns)
    existing_keys_for_display = [c for c in KEY_CANDIDATES_PRIORITY if (c in df1_raw.columns or c in df2_raw.columns)]
    core_first = uniq_list(existing_keys_for_display + ["_KEY", "_KEY_SOURCE_COL", "_SOURCE_FILE", "_SHEET"])
    candidate_keep = uniq_list(core_first + [c for c in compare_cols if c in common_cols and c not in KEY_CANDIDATES_PRIORITY])

    def _filter_dup_df(dfx: pd.DataFrame) -> pd.DataFrame:
        cols = ["Source"] + candidate_keep
        if dfx is None or dfx.empty:
            return pd.DataFrame(columns=cols)
        cols2 = ["Source"] + [c for c in candidate_keep if c in dfx.columns]
        # ✅ FIX: deduplicate column list while preserving order
        seen = set()
        cols2_dedup = []
        for c in cols2:
            if c not in seen:
                seen.add(c)
                cols2_dedup.append(c)
        return dfx[cols2_dedup].copy()

    dup_rows_1_out = _filter_dup_df(dup_rows_1)
    dup_rows_2_out = _filter_dup_df(dup_rows_2)
    dup_both_out   = _filter_dup_df(dup_both)

    # ✅ FIX: deduplicate dup outputs to ONE ROW PER KEY (for cleaner downloads)
    # Keep only first occurrence of each duplicate key in each side
    if not dup_rows_1_out.empty and "_KEY" in dup_rows_1_out.columns:
        dup_rows_1_out = dup_rows_1_out.drop_duplicates(subset=["_KEY"], keep="first").copy()
    if not dup_rows_2_out.empty and "_KEY" in dup_rows_2_out.columns:
        dup_rows_2_out = dup_rows_2_out.drop_duplicates(subset=["_KEY"], keep="first").copy()
    if not dup_both_out.empty and "_KEY" in dup_both_out.columns:
        dup_both_out = dup_both_out.drop_duplicates(subset=["_KEY", "Source"], keep="first").copy()
    
    # ============================================================
    # ✅ DUPLICATE HANDLING SECTION
    # ============================================================
    _has_dups = len(dup_union) > 0

    if _has_dups:
        st.warning(
            f"⚠️ **Duplicate rows detected** — "
            f"F1: {dup_f1_key_count:,} keys have duplicates "
            f"({len(df1_raw) - len(df1_raw.drop_duplicates(subset=['_KEY'])):,} extra rows) | "
            f"F2: {dup_f2_key_count:,} keys have duplicates "
            f"({len(df2_raw) - len(df2_raw.drop_duplicates(subset=['_KEY'])):,} extra rows)"
        )

        with st.expander("📊 Duplicate Impact Preview — click to see which keys are duplicated", expanded=False):
            dc1, dc2, dc3 = st.columns(3)
            dc1.metric("Dup keys in File1", f"{dup_f1_key_count:,}")
            dc2.metric("Dup keys in File2", f"{dup_f2_key_count:,}")
            dc3.metric("Dup keys in Union", f"{len(dup_union):,}")
            st.caption("Top duplicate keys (File1):")
            st.dataframe(dup_rows_1_out.head(20), use_container_width=True, hide_index=True)

        dup_mode = st.radio(
            "🔀 How do you want to handle duplicates in reconciliation?",
            options=[
                "✅ Include duplicates — SUM all rows per key before comparing (recommended for monthly data)",
                "🚫 Exclude duplicates — keep only FIRST row per duplicate key, remove extra rows"
            ],
            index=0,
            key="dup_mode_radio"
        )
        _exclude_dups = "Exclude" in dup_mode

        if _exclude_dups:
            _before_f1 = len(df1_raw)
            _before_f2 = len(df2_raw)

            # ✅ CORRECT APPROACH: Keep only FIRST occurrence of each duplicate key
            # This removes EXTRA duplicate rows but keeps ONE row per key
            # So if a key appears 3 times → keep 1, remove 2 extra rows
            df1_raw = df1_raw.drop_duplicates(subset=["_KEY"], keep="first").copy()
            df2_raw = df2_raw.drop_duplicates(subset=["_KEY"], keep="first").copy()

            _after_f1 = len(df1_raw)
            _after_f2 = len(df2_raw)
            _removed_f1 = _before_f1 - _after_f1
            _removed_f2 = _before_f2 - _after_f2

            st.info(
                f"🚫 Extra duplicate rows removed (kept first occurrence per key) — \n"
                f"F1: {_before_f1:,} → {_after_f1:,} rows kept | "
                f"Removed {_removed_f1:,} extra rows | \n"
                f"F2: {_before_f2:,} → {_after_f2:,} rows kept | "
                f"Removed {_removed_f2:,} extra rows"
            )

            if _removed_f1 == 0 and _removed_f2 == 0:
                st.info("✅ No extra duplicate rows found — both files already have unique keys.")
    else:
        _exclude_dups = False
        st.success("✅ No duplicates found in either file.")

    # ============================================================
    # ✅ BEFORE vs AFTER COMPARISON (only shown if dups exist)
    # Store whether this run was with or without duplicates
    # ============================================================
    st.session_state["dup_exclude_mode"] = _exclude_dups
    st.session_state["dup_union_count"]  = len(dup_union)

    # -----------------------------
    # Step 2/7: Aggregate + merge (OPTIMIZED)
    # -----------------------------
    progress_update(2, TOTAL_STEPS, "Aggregating duplicates (optimized) + building merged view...")




    
    # -----------------------------
    # Step 2/7: Aggregate + merge (OPTIMIZED)
    # -----------------------------
    progress_update(2, TOTAL_STEPS, "Aggregating duplicates (optimized) + building merged view...")

    safe_compare_cols = [c for c in compare_cols if (c in df1.columns and c in df2.columns)]
    safe_compare_cols = reorder_selected_cols(safe_compare_cols)

    if not safe_compare_cols:
        st.error("None of the selected compare columns exist in both sides.")
        st.stop()

    # Keep all compare cols + trace cols; key already in _KEY
    needed_cols = uniq_list(["_KEY", "_SOURCE_FILE", "_SHEET"] + safe_compare_cols)

    df1_need = df1_raw[[c for c in needed_cols if c in df1_raw.columns]].copy()
    df2_need = df2_raw[[c for c in needed_cols if c in df2_raw.columns]].copy()

    numeric_cols = []
    for c in safe_compare_cols:
        if c in KEY_CANDIDATES_PRIORITY:
            continue
        if is_text_by_name(c):
            continue

        n1 = pd.to_numeric(df1_need[c].where(~df1_need[c].apply(is_blank), None), errors="coerce")
        n2 = pd.to_numeric(df2_need[c].where(~df2_need[c].apply(is_blank), None), errors="coerce")
        if (n1.notna().sum() + n2.notna().sum()) > 0:
            numeric_cols.append(c)

    for c in numeric_cols:
        df1_need[c] = pd.to_numeric(df1_need[c].where(~df1_need[c].apply(is_blank), None), errors="coerce").fillna(0)
        df2_need[c] = pd.to_numeric(df2_need[c].where(~df2_need[c].apply(is_blank), None), errors="coerce").fillna(0)

    join_cols = ["_SOURCE_FILE", "_SHEET"]

    df1_ = aggregate_by_key_fast(df1_need, "_KEY", numeric_cols=numeric_cols, join_cols=join_cols)
    df2_ = aggregate_by_key_fast(df2_need, "_KEY", numeric_cols=numeric_cols, join_cols=join_cols)

    merged = df1_.merge(df2_, on="_KEY", how="outer", suffixes=("_f1", "_f2"), indicator=True)

    only_f1 = merged[merged["_merge"] == "left_only"].copy()
    only_f2 = merged[merged["_merge"] == "right_only"].copy()
    both    = merged[merged["_merge"] == "both"].copy()

    def add_trace_cols(dfx: pd.DataFrame) -> pd.DataFrame:
        dfx = dfx.copy()
        dfx["File1_Name"] = dfx.get("_SOURCE_FILE_f1", "")
        dfx["Sheet1"]     = dfx.get("_SHEET_f1", "")
        dfx["File2_Name"] = dfx.get("_SOURCE_FILE_f2", "")
        dfx["Sheet2"]     = dfx.get("_SHEET_f2", "")
        return dfx

    only_f1 = add_trace_cols(only_f1)
    only_f2 = add_trace_cols(only_f2)
    both    = add_trace_cols(both)

    # -----------------------------
    # Step 3/7: Compare by FOCUS column ONLY
    # -----------------------------
    progress_update(3, TOTAL_STEPS, "Comparing using the selected 'Particular' column only...")

    focus = reco_focus_col
    if focus in KEY_CANDIDATES_PRIORITY:
        st.error("Summary 'Particular' cannot be a key column. Please choose a different column.")
        st.stop()

    if focus not in safe_compare_cols:
        st.error(f"Selected Summary column '{focus}' is not available in both files for comparison.")
        st.stop()

    c1n = f"{focus}_f1"
    c2n = f"{focus}_f2"

    s1 = both.get(c1n)
    s2 = both.get(c2n)

    focus_is_numeric = is_numeric_column(both, focus)

    if focus_is_numeric:
        n1 = pd.to_numeric(s1, errors="coerce")
        n2 = pd.to_numeric(s2, errors="coerce")
        tol = float(numeric_tolerance)
        mismatch = (n1 - n2).abs() > (tol + 1e-9)
        if treat_blanks_as_equal:
            mismatch = mismatch & (~(s1.apply(is_blank) & s2.apply(is_blank)))
    else:
        s1s = s1.astype(str)
        s2s = s2.astype(str)
        mismatch = (s1s != s2s)
        if treat_blanks_as_equal:
            mismatch = mismatch & (~(s1.apply(is_blank) & s2.apply(is_blank)))

    both["Mismatch_Count"] = mismatch.astype(int)
    if skip_mismatch_columns:
        both["Mismatch_Columns"] = ""
    else:
        both["Mismatch_Columns"] = focus

    matched = both[both["Mismatch_Count"] == 0].copy()
    mismatched = both[both["Mismatch_Count"] > 0].copy()

    def add_selected_cols_diff_and_text_match(df: pd.DataFrame, selected_cols: list) -> pd.DataFrame:
        df = df.copy()

        for col in selected_cols:
            if col in KEY_CANDIDATES_PRIORITY:
                continue

            c1 = f"{col}_f1"
            c2 = f"{col}_f2"
            if c1 not in df.columns or c2 not in df.columns:
                continue

            col_is_numeric = is_numeric_column(df, col)
    
            if col_is_numeric:
                # ✅ Numeric: add Diff only (no Match column needed)
                n1 = pd.to_numeric(df[c1], errors="coerce")
                n2 = pd.to_numeric(df[c2], errors="coerce")
                df[f"Diff_{col}_f1-f2"] = (n1 - n2)
            else:
                # ✅ Text: add Match only (no Diff column — avoids #VALUE! in Excel)
                s1x = df[c1]
                s2x = df[c2]
                eq = s1x.astype(str) == s2x.astype(str)
                if treat_blanks_as_equal:
                    eq = eq | (s1x.apply(is_blank).astype(bool) & s2x.apply(is_blank).astype(bool))
                df[f"Match_{col}_f1=f2"] = eq

        return df

    matched    = add_selected_cols_diff_and_text_match(matched, safe_compare_cols)
    mismatched = add_selected_cols_diff_and_text_match(mismatched, safe_compare_cols)

    # -----------------------------
    # Step 4/7: Output column order
    # -----------------------------
    progress_update(4, TOTAL_STEPS, "Building Matched/Mismatched/Only-in outputs...")

    base_cols = [
        "File1_Name", "Sheet1", "File2_Name", "Sheet2",
        "_KEY"
    ]

    compare_out_cols = []
    for col in safe_compare_cols:
        if col in KEY_CANDIDATES_PRIORITY:
            continue

        c1 = f"{col}_f1"
        c2 = f"{col}_f2"
        d  = f"Diff_{col}_f1-f2"
        m  = f"Match_{col}_f1=f2"

        compare_out_cols += [c1, c2]

        # ✅ FIX: only add Diff for numeric cols, only add Match for text cols
        col_is_num = is_numeric_column(matched, col) if not matched.empty else is_numeric_column(mismatched, col) if not mismatched.empty else False
        if col_is_num:
            if d in matched.columns:
                compare_out_cols.append(d)
        else:
            if m in matched.columns:
                compare_out_cols.append(m)
    extra_cols = ["Mismatch_Count"]
    if not skip_mismatch_columns:
        extra_cols.append("Mismatch_Columns")

    # Keep desired output column order (even if some cols missing in matched/mismatched)
    keep_cols = uniq_list(base_cols + compare_out_cols + extra_cols)

    matched_out    = matched.reindex(columns=keep_cols).copy()
    mismatched_out = mismatched.reindex(columns=keep_cols).copy()

    only_cols = [
        "File1_Name", "Sheet1", "File2_Name", "Sheet2",
        "_KEY", f"{focus}_f1", f"{focus}_f2"
    ]

    diff_focus_col = f"Diff_{focus}_f1-f2"
    only_cols += [diff_focus_col]

    only_f1_out = only_f1.reindex(columns=only_cols).copy()
    only_f2_out = only_f2.reindex(columns=only_cols).copy()

    if focus_is_numeric:
        n1 = pd.to_numeric(only_f1_out.get(f"{focus}_f1"), errors="coerce").fillna(0)
        n2 = pd.to_numeric(only_f1_out.get(f"{focus}_f2"), errors="coerce").fillna(0)
        only_f1_out[diff_focus_col] = n1 - n2

        n1b = pd.to_numeric(only_f2_out.get(f"{focus}_f1"), errors="coerce").fillna(0)
        n2b = pd.to_numeric(only_f2_out.get(f"{focus}_f2"), errors="coerce").fillna(0)
        only_f2_out[diff_focus_col] = n1b - n2b
    else:
        only_f1_out[diff_focus_col] = ""
        only_f2_out[diff_focus_col] = ""

    # -----------------------------
    # Step 5/7: Summary
    # -----------------------------
    progress_update(5, TOTAL_STEPS, "Preparing summary...")

    def sum_numeric(series):
        return pd.to_numeric(series, errors="coerce").fillna(0).sum()

    focus_f1 = f"{focus}_f1"
    focus_f2 = f"{focus}_f2"

    summary_rows = [
        {
            "SNO": 1,
            "Particular": focus,
            "Count": int(len(matched_out)),
            f"{focus}_f1": float(sum_numeric(matched_out.get(focus_f1, pd.Series([], dtype="object")))),
            f"{focus}_f2": float(sum_numeric(matched_out.get(focus_f2, pd.Series([], dtype="object")))),
            "Diff": float(
                sum_numeric(matched_out.get(focus_f1, pd.Series([], dtype="object"))) -
                sum_numeric(matched_out.get(focus_f2, pd.Series([], dtype="object")))
            ),
            "Remarks": "Match"
        },
        {
            "SNO": 2,
            "Particular": focus,
            "Count": int(len(mismatched_out)),
            f"{focus}_f1": float(sum_numeric(mismatched_out.get(focus_f1, pd.Series([], dtype="object")))),
            f"{focus}_f2": float(sum_numeric(mismatched_out.get(focus_f2, pd.Series([], dtype="object")))),
            "Diff": float(
                sum_numeric(mismatched_out.get(focus_f1, pd.Series([], dtype="object"))) -
                sum_numeric(mismatched_out.get(focus_f2, pd.Series([], dtype="object")))
            ),
            "Remarks": f"Mismatch ({len(mismatched_out):,} rows)"
        },
        {
            "SNO": 3,
            "Particular": focus,
            "Count": int(len(only_f1_out)),
            f"{focus}_f1": float(sum_numeric(only_f1_out.get(focus_f1, pd.Series([], dtype="object")))),
            f"{focus}_f2": 0.0,
            "Diff": float(sum_numeric(only_f1_out.get(focus_f1, pd.Series([], dtype="object")))),
            "Remarks": "Only in File1"
        },
        {
            "SNO": 4,
            "Particular": focus,
            "Count": int(len(only_f2_out)),
            f"{focus}_f1": 0.0,
            f"{focus}_f2": float(sum_numeric(only_f2_out.get(focus_f2, pd.Series([], dtype="object")))),
            "Diff": float(0 - sum_numeric(only_f2_out.get(focus_f2, pd.Series([], dtype="object")))),
            "Remarks": "Only in File2"
        },
        {
            "SNO": 5,
            "Particular": focus,
            "Count": int(len(dup_union)),
            f"{focus}_f1": 0.0,
            f"{focus}_f2": 0.0,
            "Diff": 0.0,
            "Remarks": f"Duplicates (F1:{dup_f1_key_count:,} keys / F2:{dup_f2_key_count:,} keys) — INFO ONLY, not in Total"
        },
    ]
    summary_df = pd.DataFrame(summary_rows)

    # ✅ FIX: Exclude SNO 5 (Duplicates) from Total — duplicates are INFO ONLY
    _summary_for_total = summary_df[summary_df["SNO"].isin([1, 2, 3, 4])]
    total_row = {
        "SNO": "",
        "Particular": "Total",
        "Count": int(_summary_for_total["Count"].sum()),
        f"{focus}_f1": float(_summary_for_total[f"{focus}_f1"].sum()),
        f"{focus}_f2": float(_summary_for_total[f"{focus}_f2"].sum()),
        "Diff": float(_summary_for_total["Diff"].sum()),
        "Remarks": "(Duplicates excluded from Total)"
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    out_summary_pdf = Path("summary_report_tool1.pdf")
    build_summary_pdf(out_summary_pdf, summary_df, APP_VERSION, focus)

    out_summary_pdf = Path("summary_report_tool1.pdf")
    build_summary_pdf(out_summary_pdf, summary_df, APP_VERSION, focus)

    # ============================================================
    # ✅ BEFORE vs AFTER DUPLICATE COMPARISON TABLE
    # Only shown if duplicates exist
    # ============================================================
    if _has_dups:
        _run_label = "🚫 WITHOUT Duplicates" if _exclude_dups else "✅ WITH Duplicates (aggregated)"

        # ✅ FIX: Only save excl summary if it has meaningful data (not all zeros)
        _total_count = int(summary_df[summary_df["SNO"].isin([1,2,3,4])]["Count"].sum())

        if _exclude_dups and _total_count == 0:
            st.warning(
                "⚠️ The WITHOUT duplicates run produced 0 records — "
                "Before vs After comparison will not be shown. "
                "This means all your records have duplicate keys. "
                "Use Include mode for meaningful reconciliation."
            )
        else:
            st.session_state[f"summary_{'excl' if _exclude_dups else 'incl'}_dups"] = summary_df.copy()
            st.session_state[f"summary_{'excl' if _exclude_dups else 'incl'}_label"] = _run_label
        # Show comparison only if BOTH versions are available
        _have_incl = "summary_incl_dups" in st.session_state
        _have_excl = "summary_excl_dups" in st.session_state

        if _have_incl and _have_excl:
            st.subheader("📊 Before vs After Duplicate Removal Comparison")
            st.caption(
                "This table compares reconciliation results WITH duplicates (aggregated) "
                "vs WITHOUT duplicates (duplicate keys fully removed from both sides)."
            )

            _df_incl = st.session_state["summary_incl_dups"]
            _df_excl = st.session_state["summary_excl_dups"]

            # Build side-by-side comparison
            _focus_f1 = f"{focus}_f1"
            _focus_f2 = f"{focus}_f2"

            _comp_rows = []
            _remarks_order = ["Match", "Mismatch", "Only in File1", "Only in File2"]

            # ✅ FIX 3: Use SNO 1-4 rows only (exclude SNO 5 duplicates row)
            # This ensures WITH Dups counts are correct (Match+Mismatch+Only1+Only2)
            _df_incl_clean = _df_incl[_df_incl["SNO"].isin([1, 2, 3, 4])].copy()
            _df_excl_clean = _df_excl[_df_excl["SNO"].isin([1, 2, 3, 4])].copy()

            for _rem in _remarks_order:
                _row_incl = _df_incl_clean[_df_incl_clean["Remarks"].str.startswith(_rem, na=False)]
                _row_excl = _df_excl_clean[_df_excl_clean["Remarks"].str.startswith(_rem, na=False)]

                _cnt_incl = int(_row_incl["Count"].sum()) if not _row_incl.empty else 0
                _cnt_excl = int(_row_excl["Count"].sum()) if not _row_excl.empty else 0

                _f1_incl = float(_row_incl[_focus_f1].sum()) if (not _row_incl.empty and _focus_f1 in _row_incl.columns) else 0.0
                _f1_excl = float(_row_excl[_focus_f1].sum()) if (not _row_excl.empty and _focus_f1 in _row_excl.columns) else 0.0

                _f2_incl = float(_row_incl[_focus_f2].sum()) if (not _row_incl.empty and _focus_f2 in _row_incl.columns) else 0.0
                _f2_excl = float(_row_excl[_focus_f2].sum()) if (not _row_excl.empty and _focus_f2 in _row_excl.columns) else 0.0

                _comp_rows.append({
                    "Category": _rem,
                    "Count (WITH Dups)": _cnt_incl,
                    "Count (WITHOUT Dups)": _cnt_excl,
                    "Count Diff": _cnt_incl - _cnt_excl,
                    f"{focus}_f1 (WITH)": round(_f1_incl, 2),
                    f"{focus}_f1 (WITHOUT)": round(_f1_excl, 2),
                    f"{focus}_f2 (WITH)": round(_f2_incl, 2),
                    f"{focus}_f2 (WITHOUT)": round(_f2_excl, 2),
                })

            # ✅ FIX 2: Add Total row at bottom
            _comp_rows.append({
                "Category": "TOTAL",
                "Count (WITH Dups)": sum(r["Count (WITH Dups)"] for r in _comp_rows),
                "Count (WITHOUT Dups)": sum(r["Count (WITHOUT Dups)"] for r in _comp_rows),
                "Count Diff": sum(r["Count Diff"] for r in _comp_rows),
                f"{focus}_f1 (WITH)": round(sum(r[f"{focus}_f1 (WITH)"] for r in _comp_rows), 2),
                f"{focus}_f1 (WITHOUT)": round(sum(r[f"{focus}_f1 (WITHOUT)"] for r in _comp_rows), 2),
                f"{focus}_f2 (WITH)": round(sum(r[f"{focus}_f2 (WITH)"] for r in _comp_rows), 2),
                f"{focus}_f2 (WITHOUT)": round(sum(r[f"{focus}_f2 (WITHOUT)"] for r in _comp_rows), 2),
            })

            _comp_df = pd.DataFrame(_comp_rows)
            
            st.dataframe(_comp_df, use_container_width=True, hide_index=True)

            # Download comparison table
            _comp_csv = _comp_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇ Download Before vs After Comparison (CSV)",
                data=_comp_csv,
                file_name="reco_before_vs_after_dup_comparison.csv",
                mime="text/csv",
                key="dl_comp_csv"
            )

        elif _have_incl and not _have_excl:
            st.info(
                "💡 **Tip:** You are viewing results **WITH duplicates included**. "
                "To see the Before vs After comparison:\n\n"
                "1️⃣ Select **'🚫 Exclude duplicates'** option in the radio above\n\n"
                "2️⃣ Click **'✅ Run Reco'** button again"
            )        
        elif _have_excl and not _have_incl:
            st.info(
                "💡 **Tip:** You are viewing results WITHOUT duplicates. "
                "To see the Before vs After comparison, re-run with "
                "'✅ Include duplicates' option selected above."
            )

    key_diag_df = pd.DataFrame([{
        "Mode": "COALESCE key (no detection)",
        "KeyPriority": " / ".join(existing_key_candidates),
        "KeyCleanMode": "TEXT (alphanumeric preserved)"
    }])

    # -----------------------------
    # Step 6/7: Show results
    # -----------------------------
    progress_update(6, TOTAL_STEPS, "Rendering outputs...")

    # ✅ Show which duplicate mode is active
    _mode_label = "🚫 Excl. Dups" if _exclude_dups else "✅ Incl. Dups"
    k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
    k1.metric("Key Mode", "COALESCE")
    k2.metric("Dup Mode", _mode_label)
    k3.metric("Matched", f"{len(matched_out):,}")
    k4.metric("Mismatched", f"{len(mismatched_out):,}")
    k5.metric("Only in File1", f"{len(only_f1_out):,}")
    k6.metric("Only in File2", f"{len(only_f2_out):,}")
    k7.metric("Duplicate Keys (union)", f"{len(dup_union):,}")

    # ── UPGRADE POPUP FOR FREE USERS ──
if st.session_state.get("user_plan","free") == "free":
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0a1628,#0f2040);
                border:1px solid rgba(0,212,255,0.3); border-radius:16px;
                padding:1.5rem 2rem; margin:1rem 0; text-align:center;">
        <div style="font-size:1.1rem; font-weight:800; color:#00d4ff; margin-bottom:0.5rem;">
            🔒 You are viewing Demo Results (100 rows only)
        </div>
        <div style="color:rgba(255,255,255,0.7); font-size:0.9rem; margin-bottom:1rem;">
            Your actual files have thousands of rows. Upgrade to see complete reconciliation,
            all downloads, and PDF reports.
        </div>
        <div style="display:flex; gap:1rem; justify-content:center; flex-wrap:wrap;">
            <a href="https://aarohisharma5000.github.io/recosuite"
               target="_blank"
               style="background:#ffd600; color:#1a1a00; padding:0.7rem 2rem;
                      border-radius:50px; font-weight:800; text-decoration:none; font-size:1rem;">
                ⭐ Upgrade — ₹299 for 6 months
            </a>
            <a href="https://wa.me/919818799197?text=RecoSuite%20Subscribe"
               target="_blank"
               style="background:#25D366; color:white; padding:0.7rem 2rem;
                      border-radius:50px; font-weight:800; text-decoration:none; font-size:1rem;">
                💬 WhatsApp to Subscribe
            </a>
        </div>
    </div>
    """, unsafe_allow_html=True)
st.subheader("4) Summary")
st.dataframe(summary_df, use_container_width=True, hide_index=True)
st.subheader("5) Outputs")
tabs = st.tabs([
    "✅ Matched", "❌ Mismatched", "⬅ Only in File 1", "➡ Only in File 2",
    "🟡 Dup_File1", "🟡 Dup_File2", "🟠 Dup_Both", "🧾 Run Log"
])

with tabs[0]:
    st.dataframe(matched_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[1]:
    st.dataframe(mismatched_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[2]:
    st.dataframe(only_f1_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[3]:
    st.dataframe(only_f2_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[4]:
    st.dataframe(dup_rows_1_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[5]:
    st.dataframe(dup_rows_2_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[6]:
    st.dataframe(dup_both_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[7]:
    # ✅ Run Log tab — shows step-by-step execution log
    _log = st.session_state.get("run_log_lines", [])
    st.code("\n".join(_log) if _log else "No log available yet.", language="text")

    # -----------------------------
    # ✅ Save Run outputs to session cache (prevents rerun recompute on downloads)
    # -----------------------------
    st.session_state["run_cache_sig"] = run_inputs_sig
    st.session_state["run_cache_ready"] = True

    st.session_state["summary_df"] = summary_df
    st.session_state["key_diag_df"] = key_diag_df
    st.session_state["matched_out"] = matched_out
    st.session_state["mismatched_out"] = mismatched_out
    st.session_state["only_f1_out"] = only_f1_out
    st.session_state["only_f2_out"] = only_f2_out
    st.session_state["dup_rows_1_out"] = dup_rows_1_out
    st.session_state["dup_rows_2_out"] = dup_rows_2_out
    st.session_state["dup_both_out"] = dup_both_out
    # ✅ FIX: save dup sets so downloads work on cached runs
    st.session_state["dup_union"]        = dup_union
    st.session_state["dup_keys_1"]       = dup_keys_1
    st.session_state["dup_keys_2"]       = dup_keys_2
    st.session_state["dup_f1_key_count"] = dup_f1_key_count
    st.session_state["dup_f2_key_count"] = dup_f2_key_count
    st.session_state["_has_dups"]        = _has_dups
    st.session_state["_exclude_dups"]    = _exclude_dups
    st.session_state["miss_cols_f1"] = miss_cols_f1
    st.session_state["miss_cols_f2"] = miss_cols_f2
    st.session_state["f1_summary"] = f1_summary
    st.session_state["f2_summary"] = f2_summary

# If cached, still render the results section (same UI)
if has_cached_run:
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Key Mode", "COALESCE")
    k2.metric("Matched", f"{len(matched_out):,}")
    k3.metric("Mismatched", f"{len(mismatched_out):,}")
    k4.metric("Only in File1", f"{len(only_f1_out):,}")
    k5.metric("Only in File2", f"{len(only_f2_out):,}")
    k6.metric("Duplicate Keys (union)", f"{len(pd.Index(dup_rows_1_out.get('_KEY', pd.Series([]))).unique()):,}")
    # ── UPGRADE POPUP FOR FREE USERS ──
if st.session_state.get("user_plan","free") == "free":
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0a1628,#0f2040);
                border:1px solid rgba(0,212,255,0.3); border-radius:16px;
                padding:1.5rem 2rem; margin:1rem 0; text-align:center;">
        <div style="font-size:1.1rem; font-weight:800; color:#00d4ff; margin-bottom:0.5rem;">
            🔒 You are viewing Demo Results (100 rows only)
        </div>
        <div style="color:rgba(255,255,255,0.7); font-size:0.9rem; margin-bottom:1rem;">
            Your actual files have thousands of rows. Upgrade to see complete reconciliation,
            all downloads, and PDF reports.
        </div>
        <div style="display:flex; gap:1rem; justify-content:center; flex-wrap:wrap;">
            <a href="https://aarohisharma5000.github.io/recosuite"
               target="_blank"
               style="background:#ffd600; color:#1a1a00; padding:0.7rem 2rem;
                      border-radius:50px; font-weight:800; text-decoration:none; font-size:1rem;">
                ⭐ Upgrade — ₹299 for 6 months
            </a>
            <a href="https://wa.me/919818799197?text=RecoSuite%20Subscribe"
               target="_blank"
               style="background:#25D366; color:white; padding:0.7rem 2rem;
                      border-radius:50px; font-weight:800; text-decoration:none; font-size:1rem;">
                💬 WhatsApp to Subscribe
            </a>
        </div>
    </div>
    """, unsafe_allow_html=True)

st.subheader("4) Summary")
st.dataframe(summary_df, use_container_width=True, hide_index=True)

st.subheader("5) Outputs")
tabs = st.tabs([
    "✅ Matched", "❌ Mismatched", "⬅ Only in File 1", "➡ Only in File 2",
    "🟡 Dup_File1", "🟡 Dup_File2", "🟠 Dup_Both", "🧾 Run Log"
])
with tabs[0]:
    st.dataframe(matched_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[1]:
    st.dataframe(mismatched_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[2]:
    st.dataframe(only_f1_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[3]:
    st.dataframe(only_f2_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[4]:
    st.dataframe(dup_rows_1_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[5]:
    st.dataframe(dup_rows_2_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[6]:
    st.dataframe(dup_both_out.head(int(show_rows_in_browser)), use_container_width=True, hide_index=True)
with tabs[7]:
    _log = st.session_state.get("run_log_lines", [])
    st.code("\n".join(_log) if _log else "No log available (cached run).", language="text")

out_summary_pdf = Path("summary_report_tool1.pdf")
# -----------------------------
# Step 7/7: Download (CSV + Excel)
# -----------------------------
progress_update(7, TOTAL_STEPS, "Ready for download...")
prog.progress(100)
label.success("✅ Completed")

st.subheader("6) Download")

# ── 📦 Total Download Size Badge — covers ALL prepared files (CSV + Excel)
#    Updates live as user prepares each file
_total_bytes = 0
for _key in [
    "csv_bytes_matched", "csv_bytes_mismatched",
    "csv_bytes_only_f1", "csv_bytes_only_f2",
    "csv_bytes_zip_all",
    "excel_bytes_main", "excel_bytes_matched",
]:
    _b = st.session_state.get(_key)
    if _b is not None:
        _total_bytes += len(_b)
# Summary CSV is always generated inline — add its size too
_total_bytes += len(summary_df.to_csv(index=False).encode("utf-8"))

_total_mb = _total_bytes / (1024 * 1024)
_badge_color = "#198754" if _total_mb < 50 else ("#fd7e14" if _total_mb < 200 else "#dc3545")
st.markdown(
    f"""<div style="
        display:inline-block;
        background:{_badge_color};
        color:white;
        padding:6px 14px;
        border-radius:10px;
        font-size:13px;
        font-weight:700;
        margin-bottom:4px;
    ">📦 Total prepared download size: {_total_mb:,.2f} MB</div>""",
    unsafe_allow_html=True
)
st.caption("Size updates as you prepare each file below.")

summary_pdf_bytes = out_summary_pdf.read_bytes() if out_summary_pdf.exists() else b""

if summary_pdf_bytes:
    st.download_button(
        "⬇ Download Summary Report (PDF)",
        data=summary_pdf_bytes,
        file_name="summary_report.pdf",
        mime="application/pdf"
    )

# --------------------------------------------------
# Stable signature for this run
# --------------------------------------------------
run_sig = (
    str(len(summary_df)) + "|" +
    str(len(matched_out)) + "|" +
    str(len(mismatched_out)) + "|" +
    str(len(only_f1_out)) + "|" +
    str(len(only_f2_out)) + "|" +
    str(len(dup_rows_1_out)) + "|" +
    str(len(dup_rows_2_out)) + "|" +
    str(len(dup_both_out))
)

# ==================================================
# (A) ⚡ CSV DOWNLOADS — Grouped Expanders
# ==================================================
st.markdown("### ⚡ CSV Downloads")
st.caption("CSV is fastest for very large outputs. No formatting, only raw data.")

# Init CSV session keys
for k in [
    "csv_ready_matched","csv_bytes_matched",
    "csv_ready_mismatched","csv_bytes_mismatched",
    "csv_ready_only_f1","csv_bytes_only_f1",
    "csv_ready_only_f2","csv_bytes_only_f2",
    "csv_ready_zip_all","csv_bytes_zip_all"
]:
    if k not in st.session_state:
        st.session_state[k] = None if "bytes" in k else False

# Reset CSV ready flags if run changes
if st.session_state.get("csv_sig_last") != run_sig:
    st.session_state["csv_sig_last"] = run_sig
    st.session_state["csv_ready_matched"] = False
    st.session_state["csv_ready_mismatched"] = False
    st.session_state["csv_ready_only_f1"] = False
    st.session_state["csv_ready_only_f2"] = False
    st.session_state["csv_ready_zip_all"] = False

# ── 📋 Summary (always instant, no prepare needed)
with st.expander("📋 Summary CSV", expanded=True):
    st.caption(f"{len(summary_df):,} rows — available immediately")
    summary_csv_bytes = summary_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇ Download Summary (CSV)",
        data=summary_csv_bytes,
        file_name="reco_summary.csv",
        mime="text/csv",
        key="dl_summary_csv"
    )

# ── ✅ Matched
with st.expander(f"✅ Matched CSV  —  {len(matched_out):,} rows", expanded=False):
    if st.button("📄 Prepare Matched CSV", disabled=st.session_state["csv_ready_matched"], key="btn_prep_matched"):
        _ep = st.progress(0, text="Building Matched CSV…")
        _ep.progress(30, text="Serialising rows…")
        st.session_state["csv_bytes_matched"] = _df_to_csv_bytes(run_sig + "|CSV_MATCHED", matched_out)
        _ep.progress(100, text="✅ Done")
        st.session_state["csv_ready_matched"] = True
    if st.session_state["csv_ready_matched"] and st.session_state["csv_bytes_matched"] is not None:
        st.download_button(
            "⬇ Download Matched (CSV)",
            data=st.session_state["csv_bytes_matched"],
            file_name="reco_matched.csv",
            mime="text/csv",
            key="dl_matched_csv"
        )

# ── ❌ Mismatched
with st.expander(f"❌ Mismatched CSV  —  {len(mismatched_out):,} rows", expanded=True):
    if st.button("📄 Prepare Mismatched CSV", disabled=st.session_state["csv_ready_mismatched"], key="btn_prep_mism"):
        _ep = st.progress(0, text="Building Mismatched CSV…")
        _ep.progress(30, text="Serialising rows…")
        st.session_state["csv_bytes_mismatched"] = _df_to_csv_bytes(run_sig + "|CSV_MISMATCHED", mismatched_out)
        _ep.progress(100, text="✅ Done")
        st.session_state["csv_ready_mismatched"] = True
    if st.session_state["csv_ready_mismatched"] and st.session_state["csv_bytes_mismatched"] is not None:
        st.download_button(
            "⬇ Download Mismatched (CSV)",
            data=st.session_state["csv_bytes_mismatched"],
            file_name="reco_mismatched.csv",
            mime="text/csv",
            key="dl_mism_csv"
        )

# ── ⬅ Only in File 1
with st.expander(f"⬅ Only in File 1 CSV  —  {len(only_f1_out):,} rows", expanded=False):
    if st.button("📄 Prepare Only-in-File1 CSV", disabled=st.session_state["csv_ready_only_f1"], key="btn_prep_only1"):
        _ep = st.progress(0, text="Building Only-in-File1 CSV…")
        _ep.progress(30, text="Serialising rows…")
        st.session_state["csv_bytes_only_f1"] = _df_to_csv_bytes(run_sig + "|CSV_ONLY_F1", only_f1_out)
        _ep.progress(100, text="✅ Done")
        st.session_state["csv_ready_only_f1"] = True
    if st.session_state["csv_ready_only_f1"] and st.session_state["csv_bytes_only_f1"] is not None:
        st.download_button(
            "⬇ Download Only-in-File1 (CSV)",
            data=st.session_state["csv_bytes_only_f1"],
            file_name="reco_only_in_file1.csv",
            mime="text/csv",
            key="dl_only1_csv"
        )

# ── ➡ Only in File 2
with st.expander(f"➡ Only in File 2 CSV  —  {len(only_f2_out):,} rows", expanded=False):
    if st.button("📄 Prepare Only-in-File2 CSV", disabled=st.session_state["csv_ready_only_f2"], key="btn_prep_only2"):
        _ep = st.progress(0, text="Building Only-in-File2 CSV…")
        _ep.progress(30, text="Serialising rows…")
        st.session_state["csv_bytes_only_f2"] = _df_to_csv_bytes(run_sig + "|CSV_ONLY_F2", only_f2_out)
        _ep.progress(100, text="✅ Done")
        st.session_state["csv_ready_only_f2"] = True
    if st.session_state["csv_ready_only_f2"] and st.session_state["csv_bytes_only_f2"] is not None:
        st.download_button(
            "⬇ Download Only-in-File2 (CSV)",
            data=st.session_state["csv_bytes_only_f2"],
            file_name="reco_only_in_file2.csv",
            mime="text/csv",
            key="dl_only2_csv"
        )
# ── 🟡 Duplicates
# Init dup CSV session keys if not present
if "csv_ready_dups" not in st.session_state:
    st.session_state["csv_ready_dups"] = False
if "csv_bytes_dups" not in st.session_state:
    st.session_state["csv_bytes_dups"] = None

# Reset if run changed
if st.session_state.get("csv_sig_last") != run_sig:
    st.session_state["csv_ready_dups"] = False

with st.expander(f"🟡 Duplicates CSV  —  Unique Dup Keys: {len(dup_union):,} | F1 dup keys: {dup_f1_key_count:,} | F2 dup keys: {dup_f2_key_count:,}", expanded=False):
    st.caption("Contains duplicate keys found in File1, File2, and keys duplicated in both files.")
    if st.button("📄 Prepare Duplicates CSV", disabled=st.session_state["csv_ready_dups"], key="btn_prep_dups"):
        _ep = st.progress(0, text="Building Duplicates CSV…")
        _ep.progress(20, text="Preparing Dup File1…")
        b_dup1 = _df_to_csv_bytes(run_sig + "|CSV_DUP1", dup_rows_1_out)
        _ep.progress(50, text="Preparing Dup File2…")
        b_dup2 = _df_to_csv_bytes(run_sig + "|CSV_DUP2", dup_rows_2_out)
        _ep.progress(80, text="Preparing Dup Both…")
        b_dup_both = _df_to_csv_bytes(run_sig + "|CSV_DUP_BOTH", dup_both_out)
        _ep.progress(100, text="✅ Done")
        st.session_state["csv_bytes_dups"] = (b_dup1, b_dup2, b_dup_both)
        st.session_state["csv_ready_dups"] = True

    if st.session_state["csv_ready_dups"] and st.session_state["csv_bytes_dups"] is not None:
        b_dup1, b_dup2, b_dup_both = st.session_state["csv_bytes_dups"]
        st.download_button(
            f"⬇ Download Duplicates File1 ({len(dup_rows_1_out):,} rows)",
            data=b_dup1,
            file_name="reco_duplicates_file1.csv",
            mime="text/csv",
            key="dl_dup1_csv"
        )
        st.download_button(
            f"⬇ Download Duplicates File2 ({len(dup_rows_2_out):,} rows)",
            data=b_dup2,
            file_name="reco_duplicates_file2.csv",
            mime="text/csv",
            key="dl_dup2_csv"
        )
        st.download_button(
            f"⬇ Download Duplicates Both Files ({len(dup_both_out):,} rows)",
            data=b_dup_both,
            file_name="reco_duplicates_both.csv",
            mime="text/csv",
            key="dl_dup_both_csv"
        )
# ── 📦 ZIP — One click, all CSVs
with st.expander("📦 One-Click ZIP (All CSVs)", expanded=False):
    st.caption("Creates a single ZIP: Summary + PDF + Matched + Mismatched + Only-in File1 + Only-in File2")
    if st.button("📦 Prepare ZIP (All CSVs)", disabled=st.session_state["csv_ready_zip_all"], key="btn_prep_zip"):
        _ep = st.progress(0, text="Building ZIP…")
        _ep.progress(10, text="Preparing Matched…")
        b_matched = _df_to_csv_bytes(run_sig + "|ZIP_MATCHED", matched_out)
        _ep.progress(30, text="Preparing Mismatched…")
        b_mism    = _df_to_csv_bytes(run_sig + "|ZIP_MISM", mismatched_out)
        _ep.progress(50, text="Preparing Only-in-File1…")
        b_only1   = _df_to_csv_bytes(run_sig + "|ZIP_ONLY1", only_f1_out)
        _ep.progress(70, text="Preparing Only-in-File2…")
        b_only2   = _df_to_csv_bytes(run_sig + "|ZIP_ONLY2", only_f2_out)
        _ep.progress(85, text="Compressing into ZIP…")
        b_summary = summary_df.to_csv(index=False).encode("utf-8")

        b_dup1_zip  = _df_to_csv_bytes(run_sig + "|ZIP_DUP1", dup_rows_1_out)
        b_dup2_zip  = _df_to_csv_bytes(run_sig + "|ZIP_DUP2", dup_rows_2_out)
        b_dupboth_zip = _df_to_csv_bytes(run_sig + "|ZIP_DUP_BOTH", dup_both_out)

        zbio = io.BytesIO()
        with zipfile.ZipFile(zbio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("reco_summary.csv", b_summary)
            if summary_pdf_bytes:
                zf.writestr("summary_report.pdf", summary_pdf_bytes)
            zf.writestr("reco_matched.csv", b_matched)
            zf.writestr("reco_mismatched.csv", b_mism)
            zf.writestr("reco_only_in_file1.csv", b_only1)
            zf.writestr("reco_only_in_file2.csv", b_only2)
            # ✅ NEW: Duplicates in ZIP
            zf.writestr("reco_duplicates_file1.csv", b_dup1_zip)
            zf.writestr("reco_duplicates_file2.csv", b_dup2_zip)
            zf.writestr("reco_duplicates_both.csv", b_dupboth_zip)

        _ep.progress(100, text="✅ ZIP ready")
        st.session_state["csv_bytes_zip_all"] = zbio.getvalue()
        st.session_state["csv_ready_zip_all"] = True

    if st.session_state["csv_ready_zip_all"] and st.session_state["csv_bytes_zip_all"] is not None:
        st.download_button(
            "⬇ Download ZIP (All CSVs)",
            data=st.session_state["csv_bytes_zip_all"],
            file_name="reco_output_all_csv.zip",
            mime="application/zip",
            key="dl_zip_all"
        )

st.divider()

# ==================================================
# (B) 📊 EXCEL DOWNLOADS — Grouped Expanders
# ==================================================
st.markdown("### 📊 Excel Downloads")
st.caption("Formatted Excel with subtotals + formulas. Slightly slower than CSV.")

# Init Excel session keys
for _k, _v in [
    ("excel_bytes_main", None), ("excel_sig_main", None), ("excel_ready_main", False),
    ("excel_bytes_matched", None), ("excel_sig_matched", None), ("excel_ready_matched", False),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

@st.cache_data(show_spinner=False)
def _build_excel_cached(_sig: str, sheets: dict) -> bytes:
    return build_excel_bytes(sheets)

if st.session_state["excel_bytes_main"] is not None and st.session_state["excel_sig_main"] == run_sig:
    st.session_state["excel_ready_main"] = True
else:
    st.session_state["excel_ready_main"] = False

if st.session_state["excel_bytes_matched"] is not None and st.session_state["excel_sig_matched"] == run_sig:
    st.session_state["excel_ready_matched"] = True
else:
    st.session_state["excel_ready_matched"] = False

# ── Main Excel (without Matched)
with st.expander("✅ Main Excel — Summary + Mismatched + Only-in + Duplicates  (Recommended)", expanded=True):
    st.caption("Excludes Matched sheet for speed. Includes all other sheets with formatting.")
    prepare_main = st.button(
        "📦 Prepare Main Excel (without Matched)",
        disabled=st.session_state["excel_ready_main"],
        key="btn_prep_excel_main"
    )
    if prepare_main:
        _ep = st.progress(0, text="Building Main Excel…")
        _ep.progress(20, text="Writing sheets…")
        output_main = _build_excel_cached(run_sig + "|MAIN_NO_MATCHED", {
                "Summary": summary_df,
                "Key_Diagnostics": key_diag_df,
                "UploadedFiles_File1": f1_summary,
                "UploadedFiles_File2": f2_summary,
                "MissingCols_File1": miss_cols_f1,
                "MissingCols_File2": miss_cols_f2,
                "Mismatched": mismatched_out,
                "Only_in_File1": only_f1_out,
                "Only_in_File2": only_f2_out,
                "Duplicate_Keys_File1": dup_rows_1_out,
                "Duplicate_Keys_File2": dup_rows_2_out,
                "Duplicate_Keys_BothFiles": dup_both_out,
            })
        _ep.progress(100, text="✅ Excel ready")
        st.session_state["excel_bytes_main"] = output_main
        st.session_state["excel_sig_main"] = run_sig
        st.session_state["excel_ready_main"] = True
        st.success("✅ Main Excel is ready.")
    if st.session_state["excel_ready_main"] and st.session_state["excel_bytes_main"] is not None:
        st.download_button(
            "⬇ Download Main Reco Output (Excel)",
            data=st.session_state["excel_bytes_main"],
            file_name="reco_output_MAIN_no_matched.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_main"
        )

# ── Matched Excel (optional, slower)
with st.expander(f"📌 Matched Excel only  —  {len(matched_out):,} rows  (optional, can be slow)", expanded=False):
    st.caption("Generate only if you need Matched rows in formatted Excel.")
    prepare_matched = st.button(
        "📦 Prepare Matched Excel",
        disabled=st.session_state["excel_ready_matched"],
        key="btn_prep_excel_matched"
    )
    if prepare_matched:
        _ep = st.progress(0, text="Building Matched Excel… (may be slow for large files)")
        _ep.progress(20, text="Writing Matched sheet…")
        output_matched = _build_excel_cached(run_sig + "|MATCHED_ONLY", {
            "Matched": matched_out
        })
        _ep.progress(100, text="✅ Excel ready")
        st.session_state["excel_bytes_matched"] = output_matched
        st.session_state["excel_sig_matched"] = run_sig
        st.session_state["excel_ready_matched"] = True
        st.success("✅ Matched Excel is ready.")
    if st.session_state["excel_ready_matched"] and st.session_state["excel_bytes_matched"] is not None:
        st.download_button(
            "⬇ Download Matched (Excel)",
            data=st.session_state["excel_bytes_matched"],
            file_name="reco_output_MATCHED_only.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_matched"
        )



